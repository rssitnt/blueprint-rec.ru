from __future__ import annotations

from pathlib import Path
import shutil
from zipfile import ZipFile

import argparse
import json
import math
import re
from typing import Iterable

import requests
from openpyxl import load_workbook


API_BASE_URL = "http://127.0.0.1:8010/api"
TEST_DIR = Path(r"C:/projects/sites/blueprint-rec-2/blueprints-test")
EXPORT_DIR = Path(r"C:/projects/sites/blueprint-rec-2/exports/test-batch-2026-03-29")
IMAGE001_XLSX = Path(
    r"C:/projects/sites/blueprint-rec-2/output/current/webui_jobs/20260324_125124_42e55359/downloads/image001_coordinates.xlsx"
)
TEST1_JSON = Path(
    r"C:/projects/sites/blueprint-rec-2/output/current/runs/v2/test1_v14_filtered_plus_missing/markers_v2.json"
)
TEST1_MANUAL_EXTRAS = [
    ("1", 379, 47),
    ("29B", 368, 904),
]
POINT_TYPE = "center"
NOTE_RADIUS_RE = re.compile(r"(?:^|;)r=([0-9.]+)")
NOTE_CORE_INK_RE = re.compile(r"(?:^|;)core_ink=([0-9.]+)")
NOTE_RING_INK_RE = re.compile(r"(?:^|;)ring_ink=([0-9.]+)")
NOTE_SCORE_RE = re.compile(r"(?:^|;)post_filter_score=([0-9.]+)")

PAGE4_MARKERS = [
    ("13", 935, 446),
    ("9", 290, 591),
    ("4", 291, 714),
    ("11", 291, 774),
    ("10", 290, 817),
    ("3", 289, 916),
    ("2", 289, 1037),
    ("8", 934, 779),
    ("6", 934, 821),
    ("1", 936, 891),
    ("7", 935, 979),
    ("5", 934, 1016),
    ("14-1", 452, 1075),
    ("14-2", 512, 1097),
    ("14-3", 576, 1118),
    ("14-4(1)", 705, 1164),
    ("14-4(2)", 705, 1376),
]


def _to_marker_list(rows: Iterable[tuple[str, int | float, int | float]]) -> list[dict]:
    return [
        {"label": str(label), "x": float(x), "y": float(y), "pointType": POINT_TYPE}
        for label, x, y in rows
    ]


def load_image001_markers() -> list[dict]:
    wb = load_workbook(IMAGE001_XLSX, read_only=True)
    ws = wb.active
    markers: list[tuple[str, int, int]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[1]
        x = row[2]
        y = row[3]
        if label is None or x is None or y is None:
            continue
        markers.append((str(label), int(x), int(y)))

    wb.close()
    return _to_marker_list(markers)


def load_legacy_markers(xlsx_path: Path, scale_divisor: int = 1) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    markers: list[tuple[str, float, float]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_label = row[0] if len(row) > 0 else None
        raw_x = row[5] if len(row) > 5 else None
        raw_y = row[6] if len(row) > 6 else None
        if raw_x is None or raw_y is None:
            continue
        if not isinstance(raw_x, (int, float)) or not isinstance(raw_y, (int, float)):
            continue
        label = raw_label if raw_label is not None else raw_x
        markers.append((str(int(label)) if float(label).is_integer() else str(label), raw_x / scale_divisor, raw_y / scale_divisor))

    wb.close()
    return [
        {"label": str(label), "x": float(x), "y": float(y), "pointType": POINT_TYPE}
        for label, x, y in markers
    ]


def load_markers_v2_json(json_path: Path, point_type: str = POINT_TYPE) -> list[dict]:
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    markers = payload.get("markers", [])
    result: list[dict] = []

    for marker in markers:
        label = marker.get("label")
        point = marker.get("center") if point_type == "center" else marker.get("top_left")
        if not label or not point:
            continue
        x = point.get("x")
        y = point.get("y")
        if x is None or y is None:
            continue
        note = str(marker.get("note") or "")
        result.append(
            {
                "label": str(label),
                "x": float(x),
                "y": float(y),
                "pointType": point_type,
                "_radius": _note_float(note, NOTE_RADIUS_RE),
                "_core_ink": _note_float(note, NOTE_CORE_INK_RE),
                "_ring_ink": _note_float(note, NOTE_RING_INK_RE),
                "_score": _note_float(note, NOTE_SCORE_RE),
            }
        )

    return _resolve_close_same_label_duplicates(result)


def _note_float(note: str, pattern: re.Pattern[str]) -> float | None:
    match = pattern.search(note)
    if not match:
        return None
    try:
        return float(match.group(1))
    except ValueError:
        return None


def _marker_quality(marker: dict) -> tuple[float, float, float]:
    radius = float(marker.get("_radius") or 999.0)
    core_ink = float(marker.get("_core_ink") or 0.0)
    ring_ink = float(marker.get("_ring_ink") or 0.0)
    return (-radius, core_ink, ring_ink)


def _resolve_close_same_label_duplicates(markers: list[dict], distance_threshold: float = 45.0) -> list[dict]:
    kept: list[dict] = []
    consumed_indexes: set[int] = set()

    for idx, marker in enumerate(markers):
        if idx in consumed_indexes:
            continue

        same_cluster = [idx]
        for candidate_index in range(idx + 1, len(markers)):
            if candidate_index in consumed_indexes:
                continue
            candidate = markers[candidate_index]
            if marker["label"] != candidate["label"]:
                continue
            distance = math.hypot(float(marker["x"]) - float(candidate["x"]), float(marker["y"]) - float(candidate["y"]))
            if distance <= distance_threshold:
                same_cluster.append(candidate_index)

        if len(same_cluster) == 1:
            kept.append(_strip_marker_metrics(marker))
            continue

        best_index = max(
            same_cluster,
            key=lambda cluster_index: _marker_quality(markers[cluster_index]),
        )
        kept.append(_strip_marker_metrics(markers[best_index]))
        consumed_indexes.update(same_cluster)

    return kept


def _strip_marker_metrics(marker: dict) -> dict:
    return {
        "label": str(marker["label"]),
        "x": float(marker["x"]),
        "y": float(marker["y"]),
        "pointType": marker["pointType"],
    }


def merge_manual_markers(base_markers: list[dict], extras: list[tuple[str, int | float, int | float]]) -> list[dict]:
    merged = list(base_markers)
    existing = {
        (str(marker["label"]), round(float(marker["x"]), 3), round(float(marker["y"]), 3), marker["pointType"])
        for marker in merged
    }
    for label, x, y in extras:
        key = (str(label), round(float(x), 3), round(float(y), 3), POINT_TYPE)
        if key in existing:
            continue
        merged.append(
            {
                "label": str(label),
                "x": float(x),
                "y": float(y),
                "pointType": POINT_TYPE,
            }
        )
    return merged


def load_markers_for(file_name: str) -> list[dict]:
    if file_name == "image001.png":
        return load_image001_markers()
    if file_name == "page4.png":
        return _to_marker_list(PAGE4_MARKERS)
    if file_name == "test1.jpg":
        return merge_manual_markers(
            load_markers_v2_json(TEST1_JSON, point_type=POINT_TYPE),
            TEST1_MANUAL_EXTRAS,
        )
    raise ValueError(f"Unsupported file for marker source: {file_name}")


def _extract_session_id(payload: dict) -> str:
    session_block = payload.get("session", {})
    session_id = session_block.get("sessionId") or session_block.get("session_id")
    if not session_id:
        raise RuntimeError(f"Не удалось получить session_id из ответа API: {json.dumps(payload, ensure_ascii=False)}")
    return session_id


def api_create_session(file_name: str) -> str:
    response = requests.post(
        f"{API_BASE_URL}/sessions",
        json={"title": f"Batch export - {file_name}"},
        timeout=20,
    )
    response.raise_for_status()
    return _extract_session_id(response.json())


def api_upload_document(session_id: str, file_path: Path) -> None:
    with file_path.open("rb") as file_handle:
        response = requests.post(
            f"{API_BASE_URL}/sessions/{session_id}/document",
            files={"file": (file_path.name, file_handle, "application/octet-stream")},
            timeout=60,
        )
    response.raise_for_status()


def api_place_marker(session_id: str, marker: dict) -> None:
    response = requests.post(
        f"{API_BASE_URL}/sessions/{session_id}/commands",
        json={
            "type": "place_marker",
            "actor": "human",
            "label": marker["label"],
            "x": marker["x"],
            "y": marker["y"],
            "pointType": marker["pointType"],
        },
        timeout=20,
    )
    response.raise_for_status()


def api_export_session(session_id: str, zip_path: Path, unpack_dir: Path) -> None:
    response = requests.get(
        f"{API_BASE_URL}/sessions/{session_id}/export",
        timeout=60,
        stream=True,
    )
    response.raise_for_status()
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    if zip_path.exists():
        zip_path.unlink()
    if unpack_dir.exists():
        shutil.rmtree(unpack_dir)
    with zip_path.open("wb") as f:
        for chunk in response.iter_content(chunk_size=1024 * 1024):
            if chunk:
                f.write(chunk)

    with ZipFile(zip_path, "r") as zf:
        zf.extractall(unpack_dir)


def build_batch_exports(target_files: list[str] | None = None) -> None:
    if not target_files:
        target_files = ["image001.png", "page4.png", "test1.jpg"]
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    created_outputs: list[Path] = []

    for file_name in target_files:
        source_path = TEST_DIR / file_name
        if not source_path.exists():
            raise FileNotFoundError(f"Не найден тестовый файл: {source_path}")

        markers = load_markers_for(file_name)
        if not markers:
            print(f"[{file_name}] пусто: координаты не найдены, пропускаю")
            continue

        print(f"[{file_name}] маркеров: {len(markers)}")
        session_id = api_create_session(file_name)
        api_upload_document(session_id, source_path)

        for marker in markers:
            api_place_marker(session_id, marker)

        zip_path = EXPORT_DIR / f"{source_path.stem}_annotated.zip"
        unpack_dir = EXPORT_DIR / source_path.stem
        api_export_session(session_id, zip_path, unpack_dir)
        created_outputs.extend([zip_path, unpack_dir])
        print(f"[{file_name}] экспорт: {zip_path}")

    print("\nГотово. Итоговая директория экспорта:")
    print(EXPORT_DIR)
    print("Сформированные файлы:")
    for output in sorted(created_outputs):
        if output.is_file():
            print(f"  {output}")
        elif output.is_dir():
            print(f"  {output}/")
            for nested in sorted(output.rglob("*")):
                if nested == output:
                    continue
                if nested.is_file():
                    print(f"    {nested}")
                else:
                    print(f"    {nested}/")


def main() -> None:
    parser = argparse.ArgumentParser(description="Batch export marker overlays for test drawings.")
    parser.add_argument("--only", nargs="+", help="Rebuild only specific files from blueprints-test.")
    args = parser.parse_args()
    build_batch_exports(args.only)


if __name__ == "__main__":
    main()

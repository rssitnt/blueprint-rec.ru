from __future__ import annotations

import csv
import json
import zipfile
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models.job_schemas import DrawingJobResult, DrawingResultRow


CSV_COLUMNS = [
    "row",
    "label",
    "center_x",
    "center_y",
    "top_left_x",
    "top_left_y",
    "bbox_x",
    "bbox_y",
    "bbox_w",
    "bbox_h",
    "final_score",
    "status",
    "note",
    "source_kind",
]


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _as_csv_row(row: DrawingResultRow) -> dict[str, Any]:
    return {
        "row": row.row,
        "label": row.label,
        "center_x": row.center.x if row.center else None,
        "center_y": row.center.y if row.center else None,
        "top_left_x": row.top_left.x if row.top_left else None,
        "top_left_y": row.top_left.y if row.top_left else None,
        "bbox_x": row.bbox.x if row.bbox else None,
        "bbox_y": row.bbox.y if row.bbox else None,
        "bbox_w": row.bbox.w if row.bbox else None,
        "bbox_h": row.bbox.h if row.bbox else None,
        "final_score": row.final_score,
        "status": row.status.value,
        "note": row.note or "",
        "source_kind": row.source_kind or "",
    }


def _format_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return round(value, 4)
    return value


def _style_header(ws) -> None:
    header_fill = PatternFill("solid", fgColor="111111")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        if not values:
            continue
        max_len = min(max(len(value) for value in values) + 2, 42)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_len


def _normalize_rows(rows: Sequence[DrawingResultRow] | None, result: DrawingJobResult) -> list[DrawingResultRow]:
    if rows is None:
        return list(result.rows)
    return list(rows)


def write_result_csv(
    result: DrawingJobResult,
    out_path: str | Path,
    *,
    rows: Sequence[DrawingResultRow] | None = None,
) -> Path:
    path = Path(out_path)
    _ensure_parent(path)
    export_rows = _normalize_rows(rows, result)

    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in export_rows:
            writer.writerow({key: _format_cell(value) for key, value in _as_csv_row(row).items()})
    return path


def write_result_xlsx(
    result: DrawingJobResult,
    out_path: str | Path,
    *,
    rows: Sequence[DrawingResultRow] | None = None,
    export_label: str = "all_rows",
) -> Path:
    path = Path(out_path)
    _ensure_parent(path)
    export_rows = _normalize_rows(rows, result)

    wb = Workbook()
    ws = wb.active
    ws.title = "Coordinates"
    ws.append(CSV_COLUMNS)
    for row in export_rows:
        flat = _as_csv_row(row)
        ws.append([_format_cell(flat[column]) for column in CSV_COLUMNS])
    _style_header(ws)
    _autosize_columns(ws)

    summary = wb.create_sheet("Summary")
    summary.append(["metric", "value"])
    summary_rows = [
        ("source_file", result.source_file),
        ("source_labels_file", result.source_labels_file or ""),
        ("total_rows", result.summary.total_rows),
        ("found_count", result.summary.found_count),
        ("missing_count", result.summary.missing_count),
        ("uncertain_count", result.summary.uncertain_count),
        ("export_row_count", len(export_rows)),
        ("near_tie_ambiguity_count", result.summary.near_tie_ambiguity_count),
        ("export_mode", export_label),
        ("document_confidence", _format_cell(result.summary.document_confidence)),
        ("degraded_recognition", "yes" if result.summary.degraded_recognition else "no"),
        ("degraded_reason", result.summary.degraded_reason or ""),
        ("selected_ocr_engine", result.summary.selected_ocr_engine or ""),
        ("fallback_used", "yes" if result.summary.fallback_used else "no"),
        ("fallback_attempted", "yes" if result.summary.fallback_attempted else "no"),
        ("fallback_failure_count", result.summary.fallback_failure_count),
        ("emergency_fallback_used", "yes" if result.summary.emergency_fallback_used else "no"),
        ("emergency_fallback_reason", result.summary.emergency_fallback_reason or ""),
        ("review_recommended", "yes" if result.summary.review_recommended else "no"),
        ("status_text", result.summary.status_text),
        ("failure_message", result.summary.failure_message or ""),
    ]
    for key, value in summary_rows:
        summary.append([key, value])
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 80

    wb.save(path)
    return path


def write_result_json(result: DrawingJobResult, out_path: str | Path) -> Path:
    path = Path(out_path)
    _ensure_parent(path)
    payload = result.model_dump(mode="json", by_alias=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def write_table_csv(
    out_path: str | Path,
    *,
    columns: Sequence[str],
    rows: Sequence[Mapping[str, Any]],
) -> Path:
    path = Path(out_path)
    _ensure_parent(path)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns))
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _format_cell(row.get(column)) for column in columns})
    return path


def write_json_payload(payload: Any, out_path: str | Path) -> Path:
    path = Path(out_path)
    _ensure_parent(path)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def write_result_zip(out_path: str | Path, files: Iterable[str | Path]) -> Path:
    path = Path(out_path)
    _ensure_parent(path)

    existing_files = [Path(item) for item in files if item and Path(item).exists()]
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_path in existing_files:
            archive.write(file_path, arcname=file_path.name)
    return path

from __future__ import annotations

import csv
import asyncio
import json
import os
import re
import shutil
import subprocess
import sys
from collections import Counter
from dataclasses import dataclass
from math import cos, hypot, sin, tau
from pathlib import Path
from typing import Any, Iterable, Mapping

import pypdfium2 as pdfium
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageOps, UnidentifiedImageError

from ..core.config import settings
from ..models.job_schemas import (
    DrawingJob,
    DrawingJobResult,
    DrawingJobSummary,
    DrawingResultPage,
    DrawingResultRow,
    DrawingResultRowStatus,
    ResultBoundingBox,
    ResultPoint,
)
from ..models.schemas import CreateSessionRequest
from .result_exports import (
    write_json_payload,
    write_result_csv,
    write_result_json,
    write_result_xlsx,
    write_result_zip,
    write_table_csv,
)
from .candidate_vlm_recognizer import VisionLLMCandidateRecognizer
from .candidate_recognizer import DrawingCandidateRecognizer
from .page_vocabulary import normalize_label
from .session_store import InMemorySessionStore


LABEL_CONFIDENCE_THRESHOLD = 0.70
SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9._-]+")
NEAR_TIE_NOTE_TOKEN = "OCR near-tie ambiguity"
CANONICAL_MIN_SIDE = 1600
MAX_AUTOCONTRAST_CUTOFF = 1.0
SUFFIX_CONFUSION_EQUIVALENTS: dict[str, set[str]] = {
    "A": {"4"},
    "B": {"8"},
    "E": {"3"},
    "G": {"6"},
    "I": {"1", "L"},
    "L": {"1", "I"},
    "O": {"0", "Q", "D"},
    "Q": {"0", "O"},
    "S": {"5"},
    "Z": {"2"},
}


@dataclass(frozen=True)
class JobRunOutput:
    result: DrawingJobResult
    production_csv_path: Path
    production_xlsx_path: Path
    production_zip_path: Path
    review_csv_path: Path
    held_back_csv_path: Path | None
    near_tie_csv_path: Path | None
    near_tie_json_path: Path | None
    review_xlsx_path: Path
    review_zip_path: Path
    source_json_path: Path | None
    result_json_path: Path
    page_artifacts: list["JobPageArtifact"]


@dataclass(frozen=True)
class PreparedDrawingPage:
    page_index: int
    raster_path: Path
    width: int
    height: int


@dataclass(frozen=True)
class JobPageArtifact:
    page_index: int
    overlay_path: Path | None
    source_json_path: Path | None
    width: int
    height: int
    raster_path: Path | None = None


@dataclass(frozen=True)
class LegacyPipelineOutput:
    payload: dict[str, Any]
    report: dict[str, Any]
    markers_json_path: Path
    overlay_path: Path


def safe_slug(text: str, default: str = "drawing") -> str:
    cleaned = SAFE_NAME_RE.sub("_", str(text or "").strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("._-")
    return cleaned or default


def _normalize_ocr_engine(value: str | None) -> str:
    engine = str(value or "").strip().lower()
    return engine if engine in {"easy", "rapid", "both"} else "both"


def _fallback_ocr_engines(primary_engine: str) -> list[str]:
    primary = _normalize_ocr_engine(primary_engine)
    if primary == "both":
        return ["easy", "rapid"]
    if primary == "easy":
        return ["rapid"]
    if primary == "rapid":
        return ["easy"]
    return ["easy", "rapid"]


def _parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        try:
            return float(text.replace(",", "."))
        except ValueError:
            return None


def _normalize_marker_status(value: Any) -> str:
    text = str(value or "").strip().lower()
    if "." in text:
        text = text.rsplit(".", 1)[-1]
    return text


def _normalize_label_value(value: Any) -> str | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not value.is_integer():
            return None
        return str(int(value))
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"\d+(?:[.,]0+)?", text):
        number = int(float(text.replace(",", ".")))
        return str(number)
    return None


def _pick_best_numeric_column(rows: list[list[Any]]) -> int | None:
    if not rows:
        return None

    max_columns = min(max(len(row) for row in rows), 10)
    best_index: int | None = None
    best_score = 0
    for column_index in range(max_columns):
        score = 0
        for row in rows:
            if column_index >= len(row):
                continue
            if _normalize_label_value(row[column_index]) is not None:
                score += 1
        if score > best_score:
            best_score = score
            best_index = column_index
    return best_index if best_score > 0 else None


def _dedupe_keep_order(values: Iterable[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out


def _read_text_rows(path: Path) -> list[list[str]]:
    encodings = ("utf-8-sig", "utf-8", "cp1251", "utf-16")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                sample = handle.read(4096)
                handle.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                except csv.Error:
                    dialect = csv.excel
                return [list(row) for row in csv.reader(handle, dialect)]
        except UnicodeError as exc:
            last_error = exc
            continue
    raise RuntimeError(f"Не удалось прочитать таблицу: {path.name}") from last_error


def extract_expected_labels(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
    elif suffix in {".csv", ".txt", ".tsv"}:
        rows = _read_text_rows(path)
    else:
        raise RuntimeError(f"Неподдерживаемый формат таблицы: {path.suffix or path.name}")

    column_index = _pick_best_numeric_column(rows)
    if column_index is None:
        return []

    labels: list[str] = []
    for row in rows:
        if column_index >= len(row):
            continue
        label = _normalize_label_value(row[column_index])
        if label is not None:
            labels.append(label)
    return _dedupe_keep_order(labels)


def write_labels_xlsx(labels: list[str], out_path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Labels"
    for row_index, label in enumerate(labels, start=1):
        sheet.cell(row=row_index, column=1, value=label)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)
    return out_path


def prepare_labels_for_legacy_pipeline(labels_source: Path | None, work_dir: Path) -> tuple[Path | None, list[str]]:
    if labels_source is None:
        return None, []
    labels = extract_expected_labels(labels_source)
    if not labels:
        raise RuntimeError("Не удалось извлечь ни одного номера из таблицы.")
    labels_xlsx_path = work_dir / "labels.pipeline.xlsx"
    return write_labels_xlsx(labels, labels_xlsx_path), labels


def _canonicalize_drawing_image(image: Image.Image) -> Image.Image:
    normalized = ImageOps.exif_transpose(image)
    if normalized.mode not in {"RGB", "L"}:
        normalized = normalized.convert("RGBA" if "A" in normalized.getbands() else "RGB")
    if "A" in normalized.getbands():
        background = Image.new("RGBA", normalized.size, (255, 255, 255, 255))
        background.alpha_composite(normalized.convert("RGBA"))
        normalized = background.convert("RGB")
    else:
        normalized = normalized.convert("RGB")

    # Keep the pipeline deterministic across PDF renders and uploaded rasters.
    grayscale = ImageOps.autocontrast(
        normalized.convert("L"),
        cutoff=MAX_AUTOCONTRAST_CUTOFF,
        preserve_tone=True,
    )
    width, height = grayscale.size
    min_side = min(width, height)
    if min_side <= 0:
        raise RuntimeError("Пустое изображение после нормализации.")
    if min_side != CANONICAL_MIN_SIDE:
        scale = CANONICAL_MIN_SIDE / float(min_side)
        resized_size = (
            max(1, int(round(width * scale))),
            max(1, int(round(height * scale))),
        )
        grayscale = grayscale.resize(resized_size, Image.Resampling.LANCZOS)
    return grayscale.convert("RGB")


def _save_canonical_image(image: Image.Image, out_path: Path) -> tuple[Path, int, int]:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    canonical = _canonicalize_drawing_image(image)
    canonical.save(out_path, format="PNG")
    width, height = canonical.size
    return out_path, width, height


def normalize_image_to_png(src_path: Path, out_path: Path) -> tuple[Path, int, int]:
    try:
        with Image.open(src_path) as image:
            return _save_canonical_image(image, out_path)
    except UnidentifiedImageError as exc:
        raise RuntimeError(f"Не удалось открыть изображение: {src_path.name}") from exc


def rasterize_pdf_pages(src_path: Path, out_dir: Path, scale: float = 2.0) -> list[PreparedDrawingPage]:
    pdf = None
    try:
        pdf = pdfium.PdfDocument(str(src_path))
        if len(pdf) == 0:
            raise RuntimeError("PDF пустой.")
        prepared_pages: list[PreparedDrawingPage] = []
        out_dir.mkdir(parents=True, exist_ok=True)
        for page_index in range(len(pdf)):
            page = pdf[page_index]
            bitmap = page.render(scale=scale)
            image = bitmap.to_pil().convert("RGB")
            out_path = out_dir / f"prepared-page-{page_index + 1:03d}.png"
            out_path, width, height = _save_canonical_image(image, out_path)
            prepared_pages.append(
                PreparedDrawingPage(
                    page_index=page_index,
                    raster_path=out_path,
                    width=width,
                    height=height,
                )
            )
    except RuntimeError:
        raise
    except Exception as exc:
        raise RuntimeError(f"Не удалось прочитать PDF: {src_path.name}") from exc
    finally:
        if pdf is not None:
            pdf.close()
    return prepared_pages


def prepare_drawing_for_legacy_pipeline(source_path: Path, prepared_path: Path) -> list[PreparedDrawingPage]:
    if source_path.suffix.lower() == ".pdf":
        return rasterize_pdf_pages(source_path, prepared_path.parent / "prepared-pages")
    normalized_path, width, height = normalize_image_to_png(source_path, prepared_path)
    return [
        PreparedDrawingPage(
            page_index=0,
            raster_path=normalized_path,
            width=width,
            height=height,
        )
    ]


def _run_async_in_new_loop(awaitable):
    loop = asyncio.new_event_loop()
    try:
        asyncio.set_event_loop(loop)
        return loop.run_until_complete(awaitable)
    finally:
        try:
            loop.run_until_complete(loop.shutdown_asyncgens())
        except Exception:
            pass
        asyncio.set_event_loop(None)
        loop.close()


def _candidate_distance_sq(marker_x: float, marker_y: float, candidate: Any) -> float:
    dx = float(candidate.center_x) - marker_x
    dy = float(candidate.center_y) - marker_y
    return dx * dx + dy * dy


def _match_candidate_for_marker(marker: Any, candidates: list[Any]) -> Any | None:
    marker_label = str(getattr(marker, "label", "") or "").strip().lower()
    marker_x = float(getattr(marker, "x", 0.0) or 0.0)
    marker_y = float(getattr(marker, "y", 0.0) or 0.0)

    same_label = [
        candidate
        for candidate in candidates
        if str(getattr(candidate, "suggested_label", "") or "").strip().lower() == marker_label
    ]
    pool = same_label or candidates
    if not pool:
        return None
    return min(pool, key=lambda candidate: _candidate_distance_sq(marker_x, marker_y, candidate))


def _internal_candidate_quality(candidate: Any) -> tuple[float, float, float]:
    confidence = float(_parse_number(getattr(candidate, "suggested_confidence", None)) or 0.0)
    topology = float(_parse_number(getattr(candidate, "topology_score", None)) or 0.0)
    score = float(_parse_number(getattr(candidate, "score", None)) or 0.0)
    kind = str(getattr(getattr(candidate, "kind", None), "value", getattr(candidate, "kind", "")) or "").strip().lower()
    source = str(getattr(candidate, "suggested_source", "") or "").strip().lower()
    label = str(getattr(candidate, "suggested_label", "") or "").strip()
    evidence_bonus = 0.0
    if kind == "text":
        evidence_bonus += 0.24
    if (
        source.startswith("document-ocr:")
        or source.startswith("tile-")
        or source.startswith("vlm-locate")
        or "+circle" in source
        or "+box" in source
    ):
        evidence_bonus += 0.18
    if kind in {"circle", "box"} and not evidence_bonus:
        evidence_bonus -= 0.16
    if label.isdigit() and len(label) == 1 and kind in {"circle", "box"}:
        evidence_bonus -= 0.08
    return (confidence + evidence_bonus, topology, score)


def _build_internal_conflict_index(session: Any) -> dict[str, list[str]]:
    conflicts_by_label: dict[str, list[str]] = {}
    for conflict in getattr(session, "pipeline_conflicts", []) or []:
        label = str(getattr(conflict, "label", "") or "").strip().lower()
        message = str(getattr(conflict, "message", "") or "").strip()
        if not label or not message:
            continue
        conflicts_by_label.setdefault(label, [])
        if message not in conflicts_by_label[label]:
            conflicts_by_label[label].append(message)
    return conflicts_by_label


def _payload_from_candidate(
    *,
    index: int,
    label: str,
    candidate: Any,
    note_parts: list[str] | None = None,
) -> dict[str, Any]:
    bbox_payload = {
        "x": round(float(candidate.bbox_x), 4),
        "y": round(float(candidate.bbox_y), 4),
        "w": round(float(candidate.bbox_width), 4),
        "h": round(float(candidate.bbox_height), 4),
    }
    confidence = _parse_number(getattr(candidate, "suggested_confidence", None))
    review_status = str(getattr(getattr(candidate, "review_status", None), "value", getattr(candidate, "review_status", "")) or "")
    status = "ai_detected" if review_status == "accepted" and (confidence or 0.0) >= LABEL_CONFIDENCE_THRESHOLD else "ai_review"
    return {
        "row": index,
        "label": label,
        "center": {
            "x": round(float(candidate.center_x), 4),
            "y": round(float(candidate.center_y), 4),
        },
        "top_left": {
            "x": bbox_payload["x"],
            "y": bbox_payload["y"],
        },
        "bbox": bbox_payload,
        "final_score": confidence,
        "source_kind": str(getattr(getattr(candidate, "kind", None), "value", getattr(candidate, "kind", "")) or "").strip() or None,
        "status": status,
        "note": " ".join(note_parts or []) or None,
    }


def _build_internal_marker_payload(session: Any, *, expected_labels: list[str] | None = None) -> list[dict[str, Any]]:
    def _candidate_label_key(value: Any) -> str | None:
        normalized = _normalize_label_value(value)
        if normalized is not None:
            return normalized
        text = str(value or "").strip()
        return text or None

    def _match_expected_candidates(label: str, candidates_index: dict[str, list[Any]]) -> tuple[Any | None, str | None]:
        normalized = _normalize_label_value(label)
        if normalized is None:
            return None, None

        exact_pool = candidates_index.get(normalized, [])
        if exact_pool:
            exact_pool.sort(key=_internal_candidate_quality, reverse=True)
            return exact_pool[0], normalized

        fallback_prefixes = (f"{normalized}-", f"{normalized}(")
        fallback_pool: list[tuple[str, Any]] = []
        for candidate_label, pool in candidates_index.items():
            lower = candidate_label.lower()
            if any(lower.startswith(prefix.lower()) for prefix in fallback_prefixes):
                for candidate in pool:
                    fallback_pool.append((candidate_label, candidate))
        if not fallback_pool:
            return None, None
        fallback_pool.sort(key=lambda item: _internal_candidate_quality(item[1]), reverse=True)
        return fallback_pool[0][1], fallback_pool[0][0]

    markers_payload: list[dict[str, Any]] = []
    conflicts_by_label = _build_internal_conflict_index(session)
    candidates = list(getattr(session, "candidates", []) or [])
    normalized_expected = [_normalize_label_value(label) or str(label).strip() for label in (expected_labels or []) if str(label).strip()]

    if normalized_expected:
        candidates_by_label: dict[str, list[Any]] = {}
        for candidate in candidates:
            label_key = _candidate_label_key(getattr(candidate, "suggested_label", None))
            review_status = str(getattr(getattr(candidate, "review_status", None), "value", getattr(candidate, "review_status", "")) or "")
            if not label_key or review_status == "rejected":
                continue
            candidates_by_label.setdefault(label_key, []).append(candidate)

        for index, label in enumerate(normalized_expected, start=1):
            matched_candidate, matched_label = _match_expected_candidates(label, candidates_by_label)
            if matched_candidate is None:
                continue
            note_parts = list(conflicts_by_label.get(label.lower(), []))
            if matched_label and matched_label.lower() != label.lower():
                note_parts.append(f"Табличный номер {label} закрыт по подпозиции {matched_label}.")
            markers_payload.append(
                _payload_from_candidate(
                    index=index,
                    label=label,
                    candidate=matched_candidate,
                    note_parts=note_parts,
                )
            )
        return markers_payload

    for index, marker in enumerate(getattr(session, "markers", []) or [], start=1):
        label = str(getattr(marker, "label", "") or "").strip()
        if not label:
            continue
        matched_candidate = _match_candidate_for_marker(marker, candidates)
        bbox_payload: dict[str, float] | None = None
        source_kind: str | None = None
        suggested_confidence = None
        if matched_candidate is not None:
            bbox_payload = {
                "x": round(float(matched_candidate.bbox_x), 4),
                "y": round(float(matched_candidate.bbox_y), 4),
                "w": round(float(matched_candidate.bbox_width), 4),
                "h": round(float(matched_candidate.bbox_height), 4),
            }
            source_kind = str(getattr(matched_candidate.kind, "value", getattr(matched_candidate, "kind", "")) or "").strip() or None
            suggested_confidence = getattr(matched_candidate, "suggested_confidence", None)

        marker_confidence = getattr(marker, "confidence", None)
        final_score = _parse_number(marker_confidence)
        if final_score is None:
            final_score = _parse_number(suggested_confidence)

        note_parts = conflicts_by_label.get(label.lower(), [])
        markers_payload.append(
            {
                "row": index,
                "label": label,
                "center": {
                    "x": round(float(getattr(marker, "x", 0.0) or 0.0), 4),
                    "y": round(float(getattr(marker, "y", 0.0) or 0.0), 4),
                },
                "top_left": {
                    "x": bbox_payload["x"],
                    "y": bbox_payload["y"],
                }
                if bbox_payload is not None
                else None,
                "bbox": bbox_payload,
                "final_score": final_score,
                "source_kind": source_kind,
                "status": str(getattr(marker, "status", "") or ""),
                "note": " ".join(note_parts) if note_parts else None,
            }
        )
    return markers_payload


def run_internal_pipeline(
    *,
    image_path: Path,
    labels_xlsx_path: Path | None,
    out_dir: Path,
    log_dir: Path,
    ocr_engine: str | None = None,
) -> LegacyPipelineOutput:
    out_dir.mkdir(parents=True, exist_ok=True)
    log_dir.mkdir(parents=True, exist_ok=True)

    session_store = InMemorySessionStore()
    created = _run_async_in_new_loop(
        session_store.create_session(CreateSessionRequest(title=image_path.stem or "job-page"))
    )
    session_id = created.session.session_id

    try:
        with image_path.open("rb") as source_stream:
            upload = session_store.prepare_upload(
                session_id=session_id,
                file_name=image_path.name,
                content_type="image/png",
                source_stream=source_stream,
                size_bytes=image_path.stat().st_size,
            )

        _run_async_in_new_loop(session_store.upload_document(session_id, upload))
        annotated = _run_async_in_new_loop(session_store.auto_annotate(session_id)).session
        expected_labels = extract_expected_labels(labels_xlsx_path) if labels_xlsx_path and labels_xlsx_path.is_file() else []

        with Image.open(upload.storage_path) as source_image:
            overlay_image = session_store._render_annotated_image(source_image.convert("RGB"), annotated.markers)
            overlay_path = out_dir / "markers_v3.overlay.png"
            overlay_image.save(overlay_path, format="PNG")

        markers_payload = _build_internal_marker_payload(annotated, expected_labels=expected_labels)
        payload = {
            "source_file": image_path.name,
            "selected_ocr_engine": _normalize_ocr_engine(ocr_engine),
            "pipeline_mode": "internal-session-fallback",
            "markers": markers_payload,
            "candidate_count": len(getattr(annotated, "candidates", []) or []),
            "marker_count": len(getattr(annotated, "markers", []) or []),
            "missing_labels": list(getattr(annotated, "missing_labels", []) or []),
            "pipeline_conflict_count": len(getattr(annotated, "pipeline_conflicts", []) or []),
        }
        report = {
            "pipeline_mode": "internal-session-fallback",
            "selected_ocr_engine": _normalize_ocr_engine(ocr_engine),
            "candidate_count": len(getattr(annotated, "candidates", []) or []),
            "marker_count": len(markers_payload),
            "ai_detected_count": sum(1 for marker in getattr(annotated, "markers", []) or [] if str(getattr(marker, "status", "") or "") == "ai_detected"),
            "ai_review_count": sum(1 for marker in getattr(annotated, "markers", []) or [] if str(getattr(marker, "status", "") or "") == "ai_review"),
            "pipeline_conflict_count": len(getattr(annotated, "pipeline_conflicts", []) or []),
        }

        markers_json_path = out_dir / "markers_v3.json"
        report_path = out_dir / "run_report_v3.json"
        markers_json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        (log_dir / "pipeline.stdout.txt").write_text("internal-session-fallback\n", encoding="utf-8")
        (log_dir / "pipeline.stderr.txt").write_text("", encoding="utf-8")

        return LegacyPipelineOutput(
            payload=payload,
            report=report,
            markers_json_path=markers_json_path,
            overlay_path=overlay_path,
        )
    finally:
        try:
            _run_async_in_new_loop(session_store.delete_session(session_id))
        except Exception:
            pass


def run_legacy_pipeline(
    *,
    image_path: Path,
    labels_xlsx_path: Path | None,
    out_dir: Path,
    log_dir: Path,
    ocr_engine: str | None = None,
    timeout_seconds: int | None = None,
    extra_cli_args: list[str] | None = None,
) -> LegacyPipelineOutput:
    out_dir.mkdir(parents=True, exist_ok=True)
    log_dir.mkdir(parents=True, exist_ok=True)

    pipeline_script = Path(settings.legacy_pipeline_script)
    if not pipeline_script.is_file():
        return run_internal_pipeline(
            image_path=image_path,
            labels_xlsx_path=labels_xlsx_path,
            out_dir=out_dir,
            log_dir=log_dir,
            ocr_engine=ocr_engine,
        )

    legacy_repo = Path(settings.legacy_pipeline_repo)
    ocr_engine = _normalize_ocr_engine(ocr_engine or os.getenv("WEBUI_OCR_ENGINE", "both"))

    command = [
        sys.executable,
        str(pipeline_script),
        "--image",
        str(image_path),
        "--out-dir",
        str(out_dir),
        "--ocr-engine",
        ocr_engine,
    ]
    if labels_xlsx_path is not None:
        command.extend(["--labels-xlsx", str(labels_xlsx_path)])
    if not settings.enable_openrouter_vision or not settings.openrouter_api_key:
        command.append("--disable-gemini")
    if extra_cli_args:
        command.extend(extra_cli_args)

    env = os.environ.copy()
    env.setdefault("PYTHONIOENCODING", "utf-8")
    if settings.openrouter_api_key:
        env["OPENROUTER_API_KEY"] = settings.openrouter_api_key
    if settings.gemini_api_key:
        env["GEMINI_API_KEY"] = settings.gemini_api_key

    creationflags = 0
    if os.name == "nt":
        creationflags |= subprocess.CREATE_NO_WINDOW

    result = subprocess.run(
        command,
        cwd=str(legacy_repo),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout_seconds or settings.legacy_pipeline_timeout_seconds,
        creationflags=creationflags,
        env=env,
    )

    (log_dir / "pipeline.stdout.txt").write_text(result.stdout or "", encoding="utf-8")
    (log_dir / "pipeline.stderr.txt").write_text(result.stderr or "", encoding="utf-8")

    if result.returncode != 0:
        tail = "\n".join((result.stderr or result.stdout or "").splitlines()[-20:])
        raise RuntimeError(tail or "Legacy pipeline завершился с ошибкой.")

    markers_json_path = out_dir / "markers_v3.json"
    overlay_path = out_dir / "markers_v3.overlay.png"
    report_path = out_dir / "run_report_v3.json"
    if not markers_json_path.exists() or not overlay_path.exists():
        raise RuntimeError("Legacy pipeline отработал, но не создал итоговые файлы.")

    payload = json.loads(markers_json_path.read_text(encoding="utf-8"))
    report = {}
    if report_path.exists():
        report = json.loads(report_path.read_text(encoding="utf-8"))

    return LegacyPipelineOutput(
        payload=payload,
        report=report,
        markers_json_path=markers_json_path,
        overlay_path=overlay_path,
    )


def _marker_sort_key(marker: Mapping[str, Any]) -> tuple[Any, ...]:
    row = marker.get("row")
    page_index = int(_parse_number(marker.get("_page_index")) or 0)
    if isinstance(row, int):
        return (0, page_index, row, str(marker.get("label") or ""))

    bbox = marker.get("bbox") if isinstance(marker.get("bbox"), Mapping) else {}
    y = _parse_number(bbox.get("y")) if isinstance(bbox, Mapping) else None
    x = _parse_number(bbox.get("x")) if isinstance(bbox, Mapping) else None
    return (1, page_index, y if y is not None else 1e18, x if x is not None else 1e18, str(marker.get("label") or ""))


def _make_point(value: Mapping[str, Any] | None) -> ResultPoint | None:
    if not isinstance(value, Mapping):
        return None
    x = _parse_number(value.get("x"))
    y = _parse_number(value.get("y"))
    if x is None or y is None:
        return None
    return ResultPoint(x=x, y=y)


def _make_bbox(value: Mapping[str, Any] | None) -> ResultBoundingBox | None:
    if not isinstance(value, Mapping):
        return None
    x = _parse_number(value.get("x"))
    y = _parse_number(value.get("y"))
    w = _parse_number(value.get("w"))
    h = _parse_number(value.get("h"))
    if x is None or y is None or w is None or h is None or w <= 0 or h <= 0:
        return None
    return ResultBoundingBox(x=x, y=y, w=w, h=h)


def _score_for_marker(marker: Mapping[str, Any]) -> float | None:
    score = _parse_number(marker.get("final_score"))
    if score is None:
        return None
    status = _normalize_marker_status(marker.get("status"))
    source_kind = str(marker.get("source_kind") or "").strip().lower()
    penalty = 0.0
    if status == "ocr_only":
        penalty += 0.08
    elif status == "gemini_override":
        penalty += 0.04
    elif status == "gemini_tile_proposal":
        penalty += 0.12

    if source_kind == "free":
        penalty += 0.05
    elif source_kind == "tile_gemini":
        penalty += 0.08
    elif not source_kind:
        penalty += 0.03

    adjusted_score = score - penalty
    return max(0.0, min(1.0, adjusted_score))


def _derive_marker_note(marker: Mapping[str, Any]) -> str | None:
    parts: list[str] = []
    raw_note = str(marker.get("note") or "").strip()
    if raw_note:
        parts.append(raw_note)

    status = _normalize_marker_status(marker.get("status"))
    source_kind = str(marker.get("source_kind") or "").strip().lower()
    if status == "ocr_only":
        parts.append("Точка держится только на OCR без явного Gemini-подтверждения.")
    elif status == "gemini_override":
        parts.append("Gemini переопределил OCR; место стоит проверить вручную.")
    elif status == "gemini_tile_proposal":
        parts.append("Точка пришла из tile-предложения Gemini и считается более рискованной.")

    if source_kind == "free":
        parts.append("У точки нет явной жёсткой привязки к горизонтальному callout-контексту.")
    elif source_kind == "tile_gemini":
        parts.append("Точка найдена через tile-Gemini обход и требует более осторожной проверки.")
    if _looks_like_strong_ocr_only_without_gemini(marker):
        parts.append("Сильный OCR-only кандидат без Gemini; оставлен только для ручной проверки.")
    elif _looks_like_medium_strength_ocr_only_multidigit_without_gemini(marker):
        parts.append("Средний OCR-only кандидат без Gemini; оставлен только для ручной проверки.")

    recognition = marker.get("recognition")
    if isinstance(recognition, Mapping):
        ocr_best_label = str(recognition.get("ocr_best_label") or "").strip()
        ocr_best_score = _parse_number(recognition.get("ocr_best_score")) or 0.0
        ocr_second_score = _parse_number(recognition.get("ocr_second_score")) or 0.0
        ocr_gap = _parse_number(marker.get("ocr_gap"))
        if ocr_gap is None:
            ocr_gap = ocr_best_score - ocr_second_score
        if ocr_second_score >= 0.95 and ocr_gap < 0.02:
            parts.append(
                f"OCR почти одинаково уверен сразу в двух вариантах; текущая метка '{ocr_best_label or str(marker.get('label') or '').strip()}' неоднозначна."
            )
        final_text_sanity_reason = str(recognition.get("final_text_sanity_reason") or "").strip()
        if final_text_sanity_reason:
            parts.append(f"Legacy final-text sanity: {final_text_sanity_reason}")
        final_ink_density = _parse_number(recognition.get("final_ink_density"))
        final_dark_pixels = _parse_number(recognition.get("final_dark_pixels"))
        final_bbox_area = _parse_number(recognition.get("final_bbox_area"))
        if final_ink_density is not None and final_ink_density < 0.07:
            parts.append("У кандидата низкая плотность тёмных пикселей для нормального callout-текста.")
        if final_dark_pixels is not None and final_dark_pixels < 120:
            parts.append("У кандидата мало тёмных пикселей, текст может быть слишком слабым или служебным.")
        if final_bbox_area is not None and final_bbox_area < 220:
            parts.append("BBox очень маленький для надёжного callout-номера.")
        gemini = recognition.get("gemini")
        if isinstance(gemini, Mapping):
            gemini_status = str(gemini.get("status") or "").strip().lower()
            gemini_confidence = _parse_number(gemini.get("confidence"))
            if gemini_status == "none" and gemini_confidence is not None and gemini_confidence >= 0.85:
                parts.append("Gemini по контексту считает, что это не номер детали.")
            gemini_reason = str(gemini.get("reason") or "").strip()
            if gemini_reason:
                parts.append(f"Gemini: {gemini_reason}")

    deduped_parts: list[str] = []
    seen: set[str] = set()
    for part in parts:
        if not part or part in seen:
            continue
        seen.add(part)
        deduped_parts.append(part)
    return " ".join(deduped_parts) if deduped_parts else None


def _row_status_for_marker(marker: Mapping[str, Any]) -> DrawingResultRowStatus:
    score = _score_for_marker(marker)
    if score is None:
        return DrawingResultRowStatus.UNCERTAIN
    if score >= LABEL_CONFIDENCE_THRESHOLD:
        return DrawingResultRowStatus.FOUND
    return DrawingResultRowStatus.UNCERTAIN


def _row_from_marker(marker: Mapping[str, Any], *, row_number: int, label: str, page_index: int | None = None) -> DrawingResultRow:
    note = _derive_marker_note(marker)
    resolved_page_index = page_index
    if resolved_page_index is None:
        resolved_page_index = int(_parse_number(marker.get("_page_index")) or 0)
    return DrawingResultRow(
        row=row_number,
        label=label,
        page_index=resolved_page_index,
        center=_make_point(marker.get("center") if isinstance(marker.get("center"), Mapping) else None),
        top_left=_make_point(marker.get("top_left") if isinstance(marker.get("top_left"), Mapping) else None),
        bbox=_make_bbox(marker.get("bbox") if isinstance(marker.get("bbox"), Mapping) else None),
        final_score=_score_for_marker(marker),
        status=_row_status_for_marker(marker),
        note=note,
        source_kind=str(marker.get("source_kind") or "").strip() or None,
    )


def _missing_row(label: str, row_number: int) -> DrawingResultRow:
    return DrawingResultRow(
        row=row_number,
        label=label,
        page_index=0,
        center=None,
        top_left=None,
        bbox=None,
        final_score=None,
        status=DrawingResultRowStatus.NOT_FOUND,
        note="Не найдено автоматически",
        source_kind=None,
    )


def _looks_like_edge_furniture_without_labels(
    marker: Mapping[str, Any],
    *,
    page_width: int | None,
    page_height: int | None,
) -> bool:
    bbox = marker.get("bbox")
    if not isinstance(bbox, Mapping) or not page_width or not page_height:
        return False

    x0 = float(_parse_number(bbox.get("x")) or 0.0)
    y0 = float(_parse_number(bbox.get("y")) or 0.0)
    w = float(_parse_number(bbox.get("w")) or 0.0)
    h = float(_parse_number(bbox.get("h")) or 0.0)
    x1 = x0 + w
    y1 = y0 + h
    if w <= 0 or h <= 0:
        return False

    in_edge_band = (
        x0 <= page_width * 0.14
        or y0 <= page_height * 0.12
        or x1 >= page_width * 0.86
        or y1 >= page_height * 0.90
    )
    if not in_edge_band:
        return False

    score = _score_for_marker(marker) or 0.0
    source_kind = str(marker.get("source_kind") or "").strip().lower()
    status = _normalize_marker_status(marker.get("status"))
    label = str(marker.get("label") or "").strip()
    aspect_ratio = w / max(1.0, h)
    near_corner = (
        (x0 <= page_width * 0.18 or x1 >= page_width * 0.82)
        and (y0 <= page_height * 0.18 or y1 >= page_height * 0.82)
    )
    weak_source = source_kind in {"", "free", "tile_gemini", "tile_seed"}
    weak_status = status in {"ocr_only", "gemini_override", "gemini_tile_proposal", ""}

    if near_corner and weak_source and weak_status and score < 0.93:
        return True
    if weak_source and weak_status and aspect_ratio <= 1.35 and score < 0.9:
        return True
    if weak_source and len(label) <= 3 and score < 0.82:
        return True
    return False


def _looks_like_sparse_text_noise_without_labels(
    marker: Mapping[str, Any],
    *,
    page_width: int | None,
    page_height: int | None,
) -> bool:
    recognition = marker.get("recognition")
    if not isinstance(recognition, Mapping) or not page_width or not page_height:
        return False

    score = _score_for_marker(marker) or 0.0
    source_kind = str(marker.get("source_kind") or "").strip().lower()
    status = _normalize_marker_status(marker.get("status"))
    weak_source = source_kind in {"", "free", "tile_gemini", "tile_seed"}
    weak_status = status in {"ocr_only", "gemini_override", "gemini_tile_proposal", ""}
    if not (weak_source and weak_status):
        return False

    geometric_scale = ((max(1.0, float(page_width) * float(page_height))) ** 0.5) / 1000.0
    adaptive_scale = min(1.0, max(0.22, geometric_scale))
    min_density = 0.07
    min_dark = max(24.0, 120.0 * adaptive_scale)
    min_area = max(140.0, 550.0 * adaptive_scale)
    max_sparse_area = 10000.0

    density = _parse_number(recognition.get("final_ink_density"))
    dark_pixels = _parse_number(recognition.get("final_dark_pixels"))
    area = _parse_number(recognition.get("final_bbox_area"))
    if area is None:
        bbox = marker.get("bbox")
        if isinstance(bbox, Mapping):
            area = float(_parse_number(bbox.get("w")) or 0.0) * float(_parse_number(bbox.get("h")) or 0.0)

    if density is None or dark_pixels is None or area is None:
        return False
    if area < min_area and score < 0.88:
        return True
    if area > max_sparse_area and density < min_density and score < 0.92:
        return True
    if density < min_density and dark_pixels < min_dark and score < 0.9:
        return True
    return False


def _box_touches_page_edge(
    box: Mapping[str, Any],
    *,
    page_width: int,
    page_height: int,
    edge_ratio_x: float,
    edge_ratio_y: float,
) -> bool:
    x0 = float(_parse_number(box.get("x0")) or 0.0)
    y0 = float(_parse_number(box.get("y0")) or 0.0)
    x1 = float(_parse_number(box.get("x1")) or 0.0)
    y1 = float(_parse_number(box.get("y1")) or 0.0)
    return (
        x0 <= page_width * edge_ratio_x
        or y0 <= page_height * edge_ratio_y
        or x1 >= page_width * (1.0 - edge_ratio_x)
        or y1 >= page_height * (1.0 - edge_ratio_y)
    )


def _has_strong_positive_contextual_endorsement(marker: Mapping[str, Any]) -> bool:
    recognition = marker.get("recognition")
    if not isinstance(recognition, Mapping):
        return False

    gemini = recognition.get("gemini")
    if not isinstance(gemini, Mapping):
        return False

    gemini_status = str(gemini.get("status") or "").strip().lower()
    gemini_confidence = _parse_number(gemini.get("confidence")) or 0.0
    gemini_reason = str(gemini.get("reason") or "").strip().lower()
    if gemini_status != "ok" or gemini_confidence < 0.95 or not gemini_reason:
        return False

    positive_patterns = (
        "clear part-number callout",
        "connected to a component via a leader line",
        "valid part number",
        "part-number callout",
        "part number callout",
        "part number or item label",
        "part number or identifier",
        "part-number or identifier",
        "part label or identifier",
        "part label",
        "leader line",
    )
    if not any(pattern in gemini_reason for pattern in positive_patterns):
        return False

    source_kind = str(marker.get("source_kind") or "").strip().lower()
    status = _normalize_marker_status(marker.get("status"))
    score = _score_for_marker(marker) or 0.0
    if source_kind != "horizontal" or status != "ocr_only":
        return False
    return score >= 0.40


def _looks_like_strong_ocr_only_without_gemini(marker: Mapping[str, Any]) -> bool:
    recognition = marker.get("recognition")
    if not isinstance(recognition, Mapping):
        return False

    gemini = recognition.get("gemini")
    if isinstance(gemini, Mapping) and gemini:
        return False

    source_kind = str(marker.get("source_kind") or "").strip().lower()
    status = _normalize_marker_status(marker.get("status"))
    if source_kind != "horizontal" or status != "ocr_only":
        return False

    label_text = str(marker.get("label") or "").strip()
    if not label_text.isdigit() or len(label_text) < 2:
        return False

    raw_score = _parse_number(marker.get("final_score")) or 0.0
    if raw_score < 0.719:
        return False

    ocr_best = _parse_number(recognition.get("ocr_best_score")) or 0.0
    ocr_second = _parse_number(recognition.get("ocr_second_score")) or 0.0
    ocr_gap = _parse_number(marker.get("ocr_gap")) or 0.0
    if ocr_best < 0.999:
        return False
    if ocr_second > 0.0:
        return False

    bbox = marker.get("bbox")
    if not isinstance(bbox, Mapping):
        return False
    width = _parse_number(bbox.get("w")) or 0.0
    height = _parse_number(bbox.get("h")) or 0.0
    if width < 20 or height < 20:
        return False
    if width > 120 or height > 90:
        return False
    aspect_ratio = width / max(1.0, height)
    if aspect_ratio < 0.35 or aspect_ratio > 1.7:
        return False

    return True


def _looks_like_medium_strength_ocr_only_multidigit_without_gemini(marker: Mapping[str, Any]) -> bool:
    recognition = marker.get("recognition")
    if not isinstance(recognition, Mapping):
        return False

    gemini = recognition.get("gemini")
    if isinstance(gemini, Mapping) and gemini:
        return False

    source_kind = str(marker.get("source_kind") or "").strip().lower()
    status = _normalize_marker_status(marker.get("status"))
    if source_kind != "horizontal" or status != "ocr_only":
        return False

    label_text = str(marker.get("label") or "").strip()
    if not label_text.isdigit() or len(label_text) < 3:
        return False

    raw_score = _parse_number(marker.get("final_score")) or 0.0
    if raw_score < 0.64:
        return False

    ocr_best = _parse_number(recognition.get("ocr_best_score")) or 0.0
    ocr_second = _parse_number(recognition.get("ocr_second_score")) or 0.0
    ocr_gap = _parse_number(marker.get("ocr_gap"))
    if ocr_gap is None:
        ocr_gap = ocr_best - ocr_second

    if ocr_best < 0.99:
        return False
    if ocr_gap < 0.24:
        return False

    return True


def _looks_like_contextual_service_text_without_labels(
    marker: Mapping[str, Any],
    *,
    page_width: int | None,
    page_height: int | None,
) -> bool:
    recognition = marker.get("recognition")
    if not isinstance(recognition, Mapping) or not page_width or not page_height:
        return False

    score = _score_for_marker(marker) or 0.0
    source_kind = str(marker.get("source_kind") or "").strip().lower()
    status = _normalize_marker_status(marker.get("status"))
    weak_source = source_kind in {"", "free", "tile_gemini", "tile_seed"}
    weak_status = status in {"ocr_only", "gemini_override", "gemini_tile_proposal", ""}

    gemini = recognition.get("gemini")
    if isinstance(gemini, Mapping):
        gemini_status = str(gemini.get("status") or "").strip().lower()
        gemini_confidence = _parse_number(gemini.get("confidence")) or 0.0
        if gemini_status == "none" and gemini_confidence >= 0.9 and score < 0.94:
            return True

    if not (weak_source and weak_status):
        return False

    ocr_gap = _parse_number(marker.get("ocr_gap")) or 0.0
    crop_box = recognition.get("crop_box")
    context_box = recognition.get("context_box")
    crop_touches_edge = isinstance(crop_box, Mapping) and _box_touches_page_edge(
        crop_box,
        page_width=page_width,
        page_height=page_height,
        edge_ratio_x=0.05,
        edge_ratio_y=0.05,
    )
    context_touches_edge = isinstance(context_box, Mapping) and _box_touches_page_edge(
        context_box,
        page_width=page_width,
        page_height=page_height,
        edge_ratio_x=0.12,
        edge_ratio_y=0.12,
    )
    if crop_touches_edge and context_touches_edge and ocr_gap < 0.08 and score < 0.9:
        return True
    if context_touches_edge and ocr_gap < 0.03 and score < 0.86:
        return True
    return False


def _is_explicit_gemini_non_label_reject(marker: Mapping[str, Any]) -> bool:
    recognition = marker.get("recognition")
    if not isinstance(recognition, Mapping):
        return False

    gemini = recognition.get("gemini")
    if not isinstance(gemini, Mapping):
        return False

    gemini_reason = str(gemini.get("reason") or "").strip().lower()
    if not gemini_reason:
        return False

    explicit_non_label_patterns = (
        "company logo",
        "brand logo",
        "marketing text",
        "graphical elements",
        "no visible part number text",
        "not a valid part number",
        "multiple distinct part labels",
        "two distinct part labels",
    )
    return any(pattern in gemini_reason for pattern in explicit_non_label_patterns)


def _is_repeated_edge_tile_gemini_noise(
    marker: Mapping[str, Any],
    *,
    label_counts: Mapping[str, int],
    page_width: int | None,
    page_height: int | None,
) -> bool:
    label = str(marker.get("label") or "").strip()
    if not label or label_counts.get(label, 0) < 3:
        return False

    source_kind = str(marker.get("source_kind") or "").strip().lower()
    status = _normalize_marker_status(marker.get("status"))
    if source_kind != "tile_gemini" or status != "gemini_tile_proposal":
        return False

    if not _looks_like_edge_furniture_without_labels(marker, page_width=page_width, page_height=page_height):
        return False

    bbox = marker.get("bbox")
    if not isinstance(bbox, Mapping):
        return False
    width = _parse_number(bbox.get("w")) or 0.0
    height = _parse_number(bbox.get("h")) or 0.0
    if width < 90 or height < 90 or width > 170 or height > 170:
        return False
    aspect_ratio = width / max(1.0, height)
    if aspect_ratio < 0.8 or aspect_ratio > 1.2:
        return False

    return True


def _should_emit_marker_without_labels(
    marker: Mapping[str, Any],
    *,
    page_width: int | None = None,
    page_height: int | None = None,
) -> bool:
    score = _score_for_marker(marker)
    if score is None:
        return False

    edge_furniture = _looks_like_edge_furniture_without_labels(
        marker,
        page_width=page_width,
        page_height=page_height,
    )
    if _has_strong_positive_contextual_endorsement(marker) and not edge_furniture:
        return True
    if _looks_like_strong_ocr_only_without_gemini(marker) and not edge_furniture:
        return True
    if _looks_like_medium_strength_ocr_only_multidigit_without_gemini(marker) and not edge_furniture:
        return True
    if score < LABEL_CONFIDENCE_THRESHOLD:
        return False

    status = _normalize_marker_status(marker.get("status"))
    source_kind = str(marker.get("source_kind") or "").strip().lower()
    strong_statuses = {
        "ai_detected",
        "ocr_gemini_agree",
        "gemini_single_localize",
        "gemini_multi_verified",
        "gemini_multi_split",
    }

    if source_kind == "horizontal" and score >= LABEL_CONFIDENCE_THRESHOLD:
        return True
    if status in strong_statuses:
        return True
    if status == "ocr_only" and source_kind == "horizontal" and score >= 0.82:
        return True
    if status == "gemini_override" and source_kind in {"horizontal", ""} and score >= 0.80:
        return True
    if edge_furniture:
        return False
    if _looks_like_sparse_text_noise_without_labels(marker, page_width=page_width, page_height=page_height):
        return False
    if _looks_like_contextual_service_text_without_labels(marker, page_width=page_width, page_height=page_height):
        return False
    return False


def _held_back_row_from_marker(
    marker: Mapping[str, Any],
    *,
    row_number: int,
    label: str,
    reason: str | None = None,
) -> DrawingResultRow:
    row = _row_from_marker(marker, row_number=row_number, label=label)
    note_prefix = reason or "Кандидат удержан и не попал в итоговый список без таблицы."
    note = row.note or ""
    return row.model_copy(
        update={
            "status": DrawingResultRowStatus.UNCERTAIN,
            "note": f"{note_prefix} {note}".strip(),
        }
    )


def _has_positive_gemini_support(marker: Mapping[str, Any]) -> bool:
    status = _normalize_marker_status(marker.get("status"))
    if status in {
        "ocr_gemini_agree",
        "gemini_single_localize",
        "gemini_multi_verified",
        "gemini_multi_split",
        "gemini_override",
    }:
        return True

    recognition = marker.get("recognition")
    if not isinstance(recognition, Mapping):
        return False
    gemini = recognition.get("gemini")
    if not isinstance(gemini, Mapping):
        return False

    gemini_status = str(gemini.get("status") or "").strip().lower()
    gemini_confidence = _parse_number(gemini.get("confidence")) or 0.0
    return gemini_status == "ok" and gemini_confidence >= 0.6


def _looks_like_ocr_near_tie_ambiguity(marker: Mapping[str, Any]) -> bool:
    recognition = marker.get("recognition")
    if not isinstance(recognition, Mapping):
        return False
    ocr_best_score = _parse_number(recognition.get("ocr_best_score")) or 0.0
    ocr_second_score = _parse_number(recognition.get("ocr_second_score")) or 0.0
    ocr_gap = _parse_number(marker.get("ocr_gap"))
    if ocr_gap is None:
        ocr_gap = ocr_best_score - ocr_second_score
    if ocr_best_score < 0.98 or ocr_second_score < 0.95:
        return False
    if ocr_gap >= 0.02:
        return False
    return True


def _is_ocr_only_without_positive_gemini(marker: Mapping[str, Any]) -> bool:
    status = _normalize_marker_status(marker.get("status"))
    return status == "ocr_only" and not _has_positive_gemini_support(marker)


def _compute_degraded_recognition_reason(
    markers: list[dict[str, Any]],
    *,
    has_labels: bool,
    found_count: int,
    uncertain_count: int,
    held_back_count: int,
    document_confidence: float,
) -> str | None:
    labeled_markers = [marker for marker in markers if str(marker.get("label") or "").strip()]
    total_candidates = len(labeled_markers)
    if total_candidates < 12:
        return None

    weak_ocr_only_count = sum(1 for marker in labeled_markers if _is_ocr_only_without_positive_gemini(marker))
    positive_context_count = sum(
        1 for marker in labeled_markers if _has_positive_gemini_support(marker) or _has_strong_positive_contextual_endorsement(marker)
    )

    weak_ratio = weak_ocr_only_count / total_candidates
    positive_ratio = positive_context_count / total_candidates
    heavy_review_tail = held_back_count >= max(8, int(total_candidates * 0.25))
    weak_result_mix = uncertain_count >= max(found_count, 6)

    if weak_ratio < 0.6 or positive_ratio > 0.2:
        return None
    if not (heavy_review_tail or document_confidence < 0.6 or weak_result_mix):
        return None

    reason = (
        f"Похоже на тяжёлый {'офлайн-' if not has_labels else ''}лист: "
        f"{weak_ocr_only_count} из {total_candidates} кандидатов держатся в основном на OCR без нормального контекстного подтверждения."
    )
    if held_back_count > 0:
        reason += f" {held_back_count} слабых кандидатов система не стала ставить автоматически."
    if positive_context_count <= 2:
        reason += " Сильных контекстных подтверждений почти нет."
    return reason


def _normalize_truth_instances(
    raw_items: list[dict[str, float | str]],
    *,
    page_index: int,
    page_width: int,
    page_height: int,
) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for item in raw_items:
        label = str(item.get("label") or "").strip()
        if not label:
            continue
        try:
            x_norm = float(item.get("x"))
            y_norm = float(item.get("y"))
            confidence = float(item.get("confidence", 0.0))
        except Exception:
            continue
        if not (0.0 <= x_norm <= 1.0 and 0.0 <= y_norm <= 1.0):
            continue
        normalized.append(
            {
                "label": label,
                "normalized_label": normalize_label(label),
                "page_index": page_index,
                "x": x_norm * page_width,
                "y": y_norm * page_height,
                "confidence": max(0.0, min(1.0, confidence)),
            }
        )
    return normalized


def _build_no_table_truth_instances(
    page_artifacts: list[JobPageArtifact] | None,
) -> dict[int, list[dict[str, Any]]]:
    if not page_artifacts:
        return {}

    recognizer = VisionLLMCandidateRecognizer()
    if not recognizer.is_enabled():
        return {}

    truth_by_page: dict[int, list[dict[str, Any]]] = {}
    for artifact in page_artifacts:
        if artifact.raster_path is None or not artifact.raster_path.is_file():
            continue
        try:
            with Image.open(artifact.raster_path) as page_image:
                preview = ImageOps.exif_transpose(page_image).convert("RGB")
        except Exception:
            continue
        heavy_sheet = min(preview.size) <= CANONICAL_MIN_SIDE
        raw_items = recognizer.extract_label_instances(preview, heavy_sheet=heavy_sheet)
        if not raw_items:
            continue
        normalized = _normalize_truth_instances(
            raw_items,
            page_index=artifact.page_index,
            page_width=artifact.width,
            page_height=artifact.height,
        )
        if normalized:
            truth_by_page[artifact.page_index] = normalized
    return truth_by_page


def _truth_match_radius_px(
    marker: Mapping[str, Any],
    *,
    page_width: int | None,
    page_height: int | None,
) -> float:
    bbox = marker.get("bbox")
    width = 0.0
    height = 0.0
    if isinstance(bbox, Mapping):
        width = float(_parse_number(bbox.get("w")) or 0.0)
        height = float(_parse_number(bbox.get("h")) or 0.0)
    page_floor = 28.0
    if page_width and page_height:
        page_floor = max(page_floor, min(page_width, page_height) * 0.045)
    return max(page_floor, max(width, height) * 1.9)


def _match_marker_to_truth_instance(
    marker: Mapping[str, Any],
    truth_item: Mapping[str, Any],
    *,
    page_width: int | None,
    page_height: int | None,
) -> float | None:
    center = marker.get("center")
    if not isinstance(center, Mapping):
        return None
    marker_x = float(_parse_number(center.get("x")) or 0.0)
    marker_y = float(_parse_number(center.get("y")) or 0.0)
    truth_x = float(_parse_number(truth_item.get("x")) or 0.0)
    truth_y = float(_parse_number(truth_item.get("y")) or 0.0)
    distance = hypot(marker_x - truth_x, marker_y - truth_y)
    radius = _truth_match_radius_px(marker, page_width=page_width, page_height=page_height)
    if distance > radius:
        return None
    marker_score = _score_for_marker(marker) or 0.0
    truth_confidence = float(_parse_number(truth_item.get("confidence")) or 0.0)
    return marker_score + truth_confidence * 0.08 - (distance / max(radius, 1.0)) * 0.22


def _apply_truth_guided_no_table_selection(
    emitted_markers: list[dict[str, Any]],
    held_back_rows: list[DrawingResultRow],
    *,
    page_truth_instances: Mapping[int, list[dict[str, Any]]] | None,
    page_sizes_by_index: Mapping[int, tuple[int, int]] | None,
    page_raster_paths_by_index: Mapping[int, Path] | None,
    next_row_number: int,
) -> tuple[list[DrawingResultRow], list[DrawingResultRow], int]:
    if not emitted_markers or not page_truth_instances:
        rows = []
        for marker in emitted_markers:
            label = str(marker.get("label") or "").strip()
            row_number = int(marker.get("row") or next_row_number)
            rows.append(_row_from_marker(marker, row_number=row_number, label=label))
            next_row_number = max(next_row_number + 1, row_number + 1)
        return rows, held_back_rows, 0

    truth_labels_by_page: dict[int, set[str]] = {}
    for page_index, items in page_truth_instances.items():
        truth_labels_by_page[page_index] = {str(item.get("normalized_label") or "") for item in items if str(item.get("normalized_label") or "")}

    selected_ids: set[int] = set()
    matched_truth_keys: set[tuple[int, str, float, float]] = set()
    rows: list[DrawingResultRow] = []
    filtered_out_count = 0
    markers_by_page_and_label: dict[tuple[int, str], list[tuple[int, dict[str, Any]]]] = {}
    for marker_index, marker in enumerate(emitted_markers):
        page_index = int(_parse_number(marker.get("_page_index")) or 0)
        label_key = normalize_label(marker.get("label"))
        if not label_key:
            continue
        markers_by_page_and_label.setdefault((page_index, label_key), []).append((marker_index, marker))

    for page_index, truth_items in page_truth_instances.items():
        page_width, page_height = (page_sizes_by_index or {}).get(page_index, (None, None))
        for truth_item in sorted(truth_items, key=lambda item: float(item.get("confidence", 0.0)), reverse=True):
            label_key = str(truth_item.get("normalized_label") or "")
            candidates = [
                (marker_index, marker)
                for marker_index, marker in markers_by_page_and_label.get((page_index, label_key), [])
                if marker_index not in selected_ids
            ]
            best: tuple[int, dict[str, Any], float] | None = None
            for marker_index, marker in candidates:
                match_score = _match_marker_to_truth_instance(
                    marker,
                    truth_item,
                    page_width=page_width,
                    page_height=page_height,
                )
                if match_score is None:
                    continue
                if best is None or match_score > best[2]:
                    best = (marker_index, marker, match_score)
            if best is None:
                continue
            marker_index, marker, _ = best
            selected_ids.add(marker_index)
            matched_truth_keys.add(
                (
                    page_index,
                    label_key,
                    float(_parse_number(truth_item.get("x")) or 0.0),
                    float(_parse_number(truth_item.get("y")) or 0.0),
                )
            )
            label = str(truth_item.get("label") or marker.get("label") or "").strip()
            row_number = int(marker.get("row") or next_row_number)
            rows.append(_row_from_marker(marker, row_number=row_number, label=label))
            next_row_number = max(next_row_number + 1, row_number + 1)

    for page_index, truth_items in page_truth_instances.items():
        same_label_rows_by_page: dict[str, list[DrawingResultRow]] = {}
        for row in rows:
            if row.page_index != page_index:
                continue
            normalized = normalize_label(row.label)
            if not normalized:
                continue
            same_label_rows_by_page.setdefault(normalized, []).append(row)
        for truth_item in truth_items:
            label_key = str(truth_item.get("normalized_label") or "")
            truth_key = (
                page_index,
                label_key,
                float(_parse_number(truth_item.get("x")) or 0.0),
                float(_parse_number(truth_item.get("y")) or 0.0),
            )
            if truth_key in matched_truth_keys:
                continue
            if float(_parse_number(truth_item.get("confidence")) or 0.0) < 0.9:
                continue
            refined_truth_row = _refine_truth_only_row_from_local_crop(
                truth_item,
                raster_path=(page_raster_paths_by_index or {}).get(page_index),
                page_width=(page_sizes_by_index or {}).get(page_index, (None, None))[0],
                page_height=(page_sizes_by_index or {}).get(page_index, (None, None))[1],
                row_number=next_row_number,
                require_local_match=bool(same_label_rows_by_page.get(label_key)),
            )
            if same_label_rows_by_page.get(label_key):
                if refined_truth_row is None:
                    held_back_rows.append(
                        _truth_only_row(
                            truth_item,
                            row_number=next_row_number,
                        ).model_copy(
                            update={
                                "note": (
                                    "Кандидат удержан: full-page truth предложил ещё одну точку с той же меткой, "
                                    "но локальный crop не подтвердил её."
                                )
                            }
                        )
                    )
                    next_row_number += 1
                    continue
            rows.append(refined_truth_row or _truth_only_row(truth_item, row_number=next_row_number))
            same_label_rows_by_page.setdefault(label_key, []).append(rows[-1])
            next_row_number += 1

    for marker_index, marker in enumerate(emitted_markers):
        if marker_index in selected_ids:
            continue
        label = str(marker.get("label") or "").strip()
        if not label:
            continue
        page_index = int(_parse_number(marker.get("_page_index")) or 0)
        label_key = normalize_label(label)
        truth_labels = truth_labels_by_page.get(page_index, set())
        if label_key in truth_labels:
            filtered_out_count += 1
            held_back_rows.append(
                _held_back_row_from_marker(
                    marker,
                    row_number=int(marker.get("row") or next_row_number),
                    label=label,
                    reason="Кандидат удержан: full-page truth нашёл для этой метки другое, более точное место.",
                )
            )
            next_row_number += 1
            continue
        row_number = int(marker.get("row") or next_row_number)
        rows.append(_row_from_marker(marker, row_number=row_number, label=label))
        next_row_number = max(next_row_number + 1, row_number + 1)

    rows, held_back_rows = _suppress_redundant_truth_only_rows(
        rows,
        held_back_rows,
        page_sizes_by_index=page_sizes_by_index,
    )
    rows.sort(key=lambda row: (row.page_index, row.center.y if row.center else 0.0, row.center.x if row.center else 0.0))
    return rows, held_back_rows, filtered_out_count


def _build_rows(
    markers: list[dict[str, Any]],
    expected_labels: list[str],
    *,
    page_sizes_by_index: Mapping[int, tuple[int, int]] | None = None,
    page_truth_instances: Mapping[int, list[dict[str, Any]]] | None = None,
    page_raster_paths_by_index: Mapping[int, Path] | None = None,
) -> tuple[list[DrawingResultRow], list[DrawingResultRow], list[str], list[str], int, int]:
    sorted_markers = sorted(markers, key=_marker_sort_key)
    if not expected_labels:
        emitted_markers: list[dict[str, Any]] = []
        held_back_rows: list[DrawingResultRow] = []
        filtered_out_count = 0
        discarded_count = 0
        no_table_label_counts: Counter[str] = Counter(str(marker.get("label") or "").strip() for marker in sorted_markers if str(marker.get("label") or "").strip())
        for index, marker in enumerate(sorted_markers, start=1):
            label = str(marker.get("label") or "").strip()
            if not label:
                continue
            page_index = int(_parse_number(marker.get("_page_index")) or 0)
            page_width, page_height = (page_sizes_by_index or {}).get(page_index, (None, None))
            edge_furniture = _looks_like_edge_furniture_without_labels(
                marker,
                page_width=page_width,
                page_height=page_height,
            )
            sparse_text_noise = _looks_like_sparse_text_noise_without_labels(
                marker,
                page_width=page_width,
                page_height=page_height,
            )
            explicit_gemini_reject = _is_explicit_gemini_non_label_reject(marker)
            contextual_service_text = _looks_like_contextual_service_text_without_labels(
                marker,
                page_width=page_width,
                page_height=page_height,
            )
            if not _should_emit_marker_without_labels(marker, page_width=page_width, page_height=page_height):
                filtered_out_count += 1
                if explicit_gemini_reject:
                    discarded_count += 1
                    continue
                if _is_repeated_edge_tile_gemini_noise(
                    marker,
                    label_counts=no_table_label_counts,
                    page_width=page_width,
                    page_height=page_height,
                ):
                    discarded_count += 1
                    continue
                held_back_rows.append(
                    _held_back_row_from_marker(
                        marker,
                        row_number=index,
                        label=label,
                        reason=(
                            "Кандидат удержан у края листа как вероятная служебная цифра."
                            if edge_furniture
                            else (
                                "Кандидат удержан как слишком разреженный или слишком мелкий для надёжного callout-текста."
                                if sparse_text_noise
                                else (
                                    "Кандидат удержан по контексту: вокруг него больше похоже на служебную область, а не на номер детали."
                                    if contextual_service_text
                                    else (
                                        "Кандидат удержан как OCR near-tie ambiguity: внутри bbox есть почти равный спор между несколькими цифрами."
                                        if _looks_like_ocr_near_tie_ambiguity(marker)
                                        else None
                                    )
                                )
                            )
                        ),
                    )
                )
                continue
            emitted_markers.append(marker)
        rows, held_back_rows, truth_filtered_out = _apply_truth_guided_no_table_selection(
            emitted_markers,
            held_back_rows,
            page_truth_instances=page_truth_instances,
            page_sizes_by_index=page_sizes_by_index,
            page_raster_paths_by_index=page_raster_paths_by_index,
            next_row_number=len(sorted_markers) + 1,
        )
        filtered_out_count += truth_filtered_out
        return rows, held_back_rows, [], [], filtered_out_count, discarded_count

    def _marker_label_key(value: Any) -> str | None:
        normalized = _normalize_label_value(value)
        if normalized is not None:
            return normalized
        text = str(value or "").strip()
        return text or None

    def _base_label_variants(expected_label: str) -> set[str]:
        normalized = _normalize_label_value(expected_label)
        if normalized is None:
            return set()
        return {
            f"{normalized}-",
            f"{normalized}(",
        }

    def _match_expected_marker(expected_label: str, label_index: dict[str, dict[str, Any]]) -> tuple[dict[str, Any] | None, set[str]]:
        normalized = _normalize_label_value(expected_label)
        if normalized is None:
            return None, set()

        exact = label_index.get(normalized)
        if exact is not None:
            return exact, {normalized}

        prefixes = _base_label_variants(expected_label)
        fallback_candidates: list[tuple[str, dict[str, Any]]] = []
        for marker_label, marker in label_index.items():
            lower = marker_label.lower()
            if any(lower.startswith(prefix.lower()) for prefix in prefixes):
                fallback_candidates.append((marker_label, marker))

        if not fallback_candidates:
            return None, set()

        fallback_candidates.sort(key=lambda item: (_score_for_marker(item[1]) or 0.0), reverse=True)
        chosen_label, chosen_marker = fallback_candidates[0]
        matched_labels = {label for label, _ in fallback_candidates}
        marker_note = str(chosen_marker.get("note") or "").strip()
        fallback_note = f"Табличный номер {expected_label} закрыт по подпозиции {chosen_label}."
        chosen_marker = dict(chosen_marker)
        chosen_marker["note"] = f"{marker_note} {fallback_note}".strip() if marker_note else fallback_note
        return chosen_marker, matched_labels

    best_by_label: dict[str, dict[str, Any]] = {}
    for marker in sorted_markers:
        label = _marker_label_key(marker.get("label"))
        if label is None:
            continue
        existing = best_by_label.get(label)
        if existing is None or (_score_for_marker(marker) or 0.0) > (_score_for_marker(existing) or 0.0):
            best_by_label[label] = marker

    rows: list[DrawingResultRow] = []
    missing_labels: list[str] = []
    normalized_expected = {_marker_label_key(label): label for label in expected_labels}
    consumed_marker_labels: set[str] = set()
    for index, original_label in enumerate(expected_labels, start=1):
        marker, matched_labels = _match_expected_marker(original_label, best_by_label)
        if marker is None:
            rows.append(_missing_row(original_label, index))
            missing_labels.append(original_label)
            continue
        consumed_marker_labels.update(matched_labels)
        rows.append(_row_from_marker(marker, row_number=index, label=original_label))

    extra_detected_labels = [
        label
        for label in _dedupe_keep_order(
            str(marker.get("label") or "").strip()
            for marker in sorted_markers
            if (
                _marker_label_key(marker.get("label")) not in normalized_expected
                and _marker_label_key(marker.get("label")) not in consumed_marker_labels
            )
        )
        if label
    ]
    return rows, [], missing_labels, extra_detected_labels, 0, 0


def _compute_summary(
    rows: list[DrawingResultRow],
    *,
    markers: list[dict[str, Any]],
    held_back_rows: list[DrawingResultRow],
    has_labels: bool,
    total_pages: int,
    filtered_out_count: int,
    discarded_count: int,
    selected_ocr_engine: str | None = None,
    fallback_used: bool = False,
) -> DrawingJobSummary:
    total_rows = len(rows)
    found_count = sum(1 for row in rows if row.status == DrawingResultRowStatus.FOUND)
    missing_count = sum(1 for row in rows if row.status == DrawingResultRowStatus.NOT_FOUND)
    uncertain_count = sum(1 for row in rows if row.status == DrawingResultRowStatus.UNCERTAIN)
    weak_context_count = sum(1 for row in rows if row.source_kind in {None, "free", "tile_gemini"})

    scores = [row.final_score for row in rows if row.final_score is not None]
    average_score = (sum(scores) / len(scores)) if scores else 0.0
    coverage = (found_count / total_rows) if total_rows else 0.0
    uncertainty_ratio = (uncertain_count / total_rows) if total_rows else 0.0
    confidence = average_score * (0.5 + 0.5 * coverage) * (1.0 - 0.5 * uncertainty_ratio)
    if total_rows == 0:
        confidence = 0.0
    confidence = max(0.0, min(1.0, confidence))

    found_or_uncertain_pages = {row.page_index for row in rows if row.status != DrawingResultRowStatus.NOT_FOUND}
    empty_pages = max(total_pages - len(found_or_uncertain_pages), 0)
    review_reasons: list[str] = []

    failure_message: str | None = None
    if found_count == 0:
        failure_message = "Система не смогла уверенно найти ни одной точки на чертеже."
    elif has_labels and total_rows and coverage < 0.5:
        failure_message = "Система не смогла надежно разобрать значительную часть номеров из таблицы."

    review_recommended = confidence < 0.85 or missing_count > 0 or uncertain_count > 0
    if has_labels and missing_count > 0:
        review_reasons.append(f"Не найдены {missing_count} номеров из таблицы.")
    if uncertain_count > 0:
        review_reasons.append(f"{uncertain_count} точек ниже порога уверенности 70%.")
    if empty_pages > 0:
        review_reasons.append(f"На {empty_pages} стр. система не нашла ни одной подходящей точки.")
    if weak_context_count > 0:
        review_reasons.append(f"{weak_context_count} точек пришли из более слабого контекста и требуют ручной проверки.")
    near_tie_held_back_count = sum(1 for row in held_back_rows if row.note and NEAR_TIE_NOTE_TOKEN in row.note)
    held_back_count = max(filtered_out_count - discarded_count, 0)
    degraded_reason = _compute_degraded_recognition_reason(
        markers,
        has_labels=has_labels,
        found_count=found_count,
        uncertain_count=uncertain_count,
        held_back_count=held_back_count,
        document_confidence=confidence,
    )
    if held_back_count > 0:
        review_reasons.append(f"{held_back_count} слабых кандидатов удержаны для ручной проверки и не попали в итоговый список.")
    if near_tie_held_back_count > 0:
        review_reasons.append(
            f"{near_tie_held_back_count} удержанных кандидатов выглядят как OCR near-tie ambiguity: внутри bbox почти равный спор между несколькими цифрами."
        )
    if discarded_count > 0:
        review_reasons.append(f"{discarded_count} кандидатов отброшены автоматически как явный служебный мусор.")
    if degraded_reason:
        review_reasons.append(degraded_reason)
    if not has_labels and confidence < 0.85 and found_count > 0:
        review_reasons.append("Распознавание шло без таблицы, поэтому часть найденных номеров стоит проверить вручную.")
    if total_rows == 0 and not review_reasons:
        review_reasons.append("Система не нашла ни одной подходящей точки.")

    if failure_message:
        status_text = failure_message
    elif degraded_reason:
        status_text = "Похоже на тяжёлый документ со слабым OCR-контуром. Нужна ручная проверка результата."
    elif review_recommended:
        if has_labels:
            status_text = f"Найдены не все точки: {found_count} из {total_rows}. Нужна проверка."
        else:
            status_text = f"Есть спорные результаты: найдено {found_count}, неуверенных {uncertain_count}. Нужна проверка."
    else:
        status_text = f"Готово: найдено {found_count} точек."

    return DrawingJobSummary(
        total_rows=total_rows,
        found_count=found_count,
        missing_count=missing_count,
        uncertain_count=uncertain_count,
        held_back_count=held_back_count,
        near_tie_ambiguity_count=near_tie_held_back_count,
        discarded_count=discarded_count,
        document_confidence=round(confidence, 4) if total_rows else 0.0,
        degraded_recognition=degraded_reason is not None,
        degraded_reason=degraded_reason,
        selected_ocr_engine=selected_ocr_engine,
        fallback_used=fallback_used,
        review_recommended=review_recommended,
        review_reasons=review_reasons,
        status_text=status_text,
        failure_message=failure_message,
    )


def build_result_from_legacy_output(
    *,
    job: DrawingJob,
    page_artifacts: list[JobPageArtifact],
    page_payloads: list[dict[str, Any]],
    expected_labels: list[str],
    selected_ocr_engine: str | None = None,
    fallback_used: bool = False,
) -> DrawingJobResult:
    all_markers: list[dict[str, Any]] = []
    for artifact, payload in zip(page_artifacts, page_payloads, strict=False):
        for item in payload.get("markers", []):
            if not isinstance(item, Mapping):
                continue
            marker = dict(item)
            marker["_page_index"] = artifact.page_index
            all_markers.append(marker)

    page_sizes_by_index = {artifact.page_index: (artifact.width, artifact.height) for artifact in page_artifacts}
    page_raster_paths_by_index = {
        artifact.page_index: artifact.raster_path
        for artifact in page_artifacts
        if artifact.raster_path is not None
    }
    page_truth_instances = _build_no_table_truth_instances(page_artifacts) if not expected_labels else {}
    rows, held_back_rows, missing_labels, extra_detected_labels, filtered_out_count, discarded_count = _build_rows(
        all_markers,
        expected_labels,
        page_sizes_by_index=page_sizes_by_index,
        page_truth_instances=page_truth_instances,
        page_raster_paths_by_index=page_raster_paths_by_index,
    )
    summary = _compute_summary(
        rows,
        markers=all_markers,
        held_back_rows=held_back_rows,
        has_labels=job.input.has_labels,
        total_pages=len(page_artifacts),
        filtered_out_count=filtered_out_count,
        discarded_count=discarded_count,
        selected_ocr_engine=selected_ocr_engine,
        fallback_used=fallback_used,
    )
    row_count_by_page: dict[int, int] = {}
    held_back_count_by_page: dict[int, int] = {}
    for row in rows:
        if row.status == DrawingResultRowStatus.NOT_FOUND:
            continue
        row_count_by_page[row.page_index] = row_count_by_page.get(row.page_index, 0) + 1
    for row in held_back_rows:
        held_back_count_by_page[row.page_index] = held_back_count_by_page.get(row.page_index, 0) + 1

    pages = [
        DrawingResultPage(
            page_index=artifact.page_index,
            width=artifact.width,
            height=artifact.height,
            overlay_url=None,
            row_count=row_count_by_page.get(artifact.page_index, 0),
            held_back_count=held_back_count_by_page.get(artifact.page_index, 0),
        )
        for artifact in page_artifacts
    ]

    return DrawingJobResult(
        source_file=job.input.drawing_name,
        source_labels_file=job.input.labels_name,
        pages=pages,
        rows=rows,
        held_back_rows=held_back_rows,
        missing_labels=missing_labels,
        extra_detected_labels=extra_detected_labels,
        summary=summary,
    )


def _truth_only_row(truth_item: Mapping[str, Any], *, row_number: int) -> DrawingResultRow:
    x = float(_parse_number(truth_item.get("x")) or 0.0)
    y = float(_parse_number(truth_item.get("y")) or 0.0)
    confidence = float(_parse_number(truth_item.get("confidence")) or 0.0)
    page_index = int(_parse_number(truth_item.get("page_index")) or 0)
    label = str(truth_item.get("label") or "").strip()
    point = ResultPoint(x=x, y=y)
    return DrawingResultRow(
        row=row_number,
        label=label,
        page_index=page_index,
        center=point,
        top_left=point,
        bbox=None,
        final_score=max(0.0, min(1.0, confidence)),
        status=DrawingResultRowStatus.UNCERTAIN,
        note="Точка поставлена по full-page truth, локальный детектор не дал надёжного центра.",
        source_kind="vlm_locator",
    )


def _truth_only_duplicate_radius_px(page_width: int | None, page_height: int | None) -> float:
    base = 96.0
    if page_width and page_height:
        base = max(base, min(page_width, page_height) * 0.08)
    return base


def _targeted_label_match_score(candidate_label: str, target_label: str) -> float:
    candidate = str(candidate_label or "").strip().upper().translate(str.maketrans({"А": "A", "В": "B", "Е": "E", "К": "K", "М": "M", "Н": "H", "О": "O", "Р": "P", "С": "C", "Т": "T", "У": "Y", "Х": "X"}))
    target = str(target_label or "").strip().upper().translate(str.maketrans({"А": "A", "В": "B", "Е": "E", "К": "K", "М": "M", "Н": "H", "О": "O", "Р": "P", "С": "C", "Т": "T", "У": "Y", "Х": "X"}))
    if not candidate or not target:
        return 0.0
    if candidate == target:
        return 1.0
    if len(candidate) != len(target):
        return 0.0
    if len(target) < 2:
        return 0.0
    if candidate[:-1] != target[:-1]:
        return 0.0
    target_suffix = target[-1]
    candidate_suffix = candidate[-1]
    if candidate_suffix == target_suffix:
        return 1.0
    equivalents = SUFFIX_CONFUSION_EQUIVALENTS.get(target_suffix, set())
    if candidate_suffix in equivalents:
        return 0.93
    reverse_equivalents = SUFFIX_CONFUSION_EQUIVALENTS.get(candidate_suffix, set())
    if target_suffix in reverse_equivalents:
        return 0.88
    return 0.0


def _mean_darkness(values: list[int]) -> float:
    if not values:
        return 0.0
    return sum(255 - value for value in values) / (len(values) * 255.0)


def _sample_circle_pixels(gray_image: Image.Image, center_x: float, center_y: float, radius: int, *, samples: int = 64) -> list[int]:
    pixels: list[int] = []
    if radius <= 0:
        return pixels
    for index in range(samples):
        angle = tau * index / float(samples)
        sample_x = int(round(center_x + cos(angle) * radius))
        sample_y = int(round(center_y + sin(angle) * radius))
        if 0 <= sample_x < gray_image.width and 0 <= sample_y < gray_image.height:
            pixels.append(int(gray_image.getpixel((sample_x, sample_y))))
    return pixels


def _sample_disk_pixels(gray_image: Image.Image, center_x: float, center_y: float, radius: int, *, step: int = 2) -> list[int]:
    pixels: list[int] = []
    if radius <= 0:
        return pixels
    radius_sq = radius * radius
    for dx in range(-radius, radius + 1, step):
        for dy in range(-radius, radius + 1, step):
            if dx * dx + dy * dy > radius_sq:
                continue
            sample_x = int(round(center_x + dx))
            sample_y = int(round(center_y + dy))
            if 0 <= sample_x < gray_image.width and 0 <= sample_y < gray_image.height:
                pixels.append(int(gray_image.getpixel((sample_x, sample_y))))
    return pixels


def _find_circle_callout_anchor(
    image: Image.Image,
    center_x: float,
    center_y: float,
) -> tuple[float, float, int, float] | None:
    gray_image = ImageOps.grayscale(image)
    best_candidate: tuple[float, float, int, float, float, float] | None = None

    for offset_x in range(-12, 13, 2):
        for offset_y in range(-12, 13, 2):
            candidate_x = center_x + offset_x
            candidate_y = center_y + offset_y
            for radius in range(12, 34, 2):
                ring_pixels = _sample_circle_pixels(gray_image, candidate_x, candidate_y, radius)
                outer_pixels = _sample_circle_pixels(gray_image, candidate_x, candidate_y, radius + 5)
                inner_pixels = _sample_disk_pixels(gray_image, candidate_x, candidate_y, max(3, int(round(radius * 0.45))))
                if len(ring_pixels) < 40 or not inner_pixels or not outer_pixels:
                    continue

                ring_darkness = _mean_darkness(ring_pixels)
                outer_darkness = _mean_darkness(outer_pixels)
                inner_darkness = _mean_darkness(inner_pixels)
                dark_fraction = sum(1 for value in ring_pixels if value < 180) / float(len(ring_pixels))
                score = (
                    ring_darkness * 1.4
                    + dark_fraction * 0.9
                    - inner_darkness * 0.25
                    - outer_darkness * 0.35
                )
                if best_candidate is None or score > best_candidate[3]:
                    best_candidate = (
                        candidate_x,
                        candidate_y,
                        radius,
                        score,
                        ring_darkness,
                        dark_fraction,
                    )

    if best_candidate is None:
        return None
    candidate_x, candidate_y, radius, score, ring_darkness, dark_fraction = best_candidate
    if score < 0.82 or ring_darkness < 0.42 or dark_fraction < 0.52:
        return None
    return candidate_x, candidate_y, radius, score


def _refine_truth_only_row_from_circle_crop(
    truth_item: Mapping[str, Any],
    *,
    image: Image.Image,
    label: str,
    normalized_label: str,
    center_x: float,
    center_y: float,
    radius: int,
    row_number: int,
    min_side: int,
    local_recognizer: DrawingCandidateRecognizer,
) -> DrawingResultRow | None:
    crop_padding = max(12, int(round(radius * 0.85)))
    left = max(0, int(round(center_x - radius - crop_padding)))
    top = max(0, int(round(center_y - radius - crop_padding)))
    right = min(image.width, int(round(center_x + radius + crop_padding)))
    bottom = min(image.height, int(round(center_y + radius + crop_padding)))
    if right - left < 24 or bottom - top < 24:
        return None

    crop = image.crop((left, top, right, bottom))
    local = local_recognizer.recognize(crop, "text")
    if normalize_label(local.label) == normalized_label and (local.confidence or 0.0) >= 0.66:
        return _truth_only_row(truth_item, row_number=row_number).model_copy(
            update={
                "center": ResultPoint(x=round(center_x, 2), y=round(center_y, 2)),
                "top_left": ResultPoint(x=round(center_x, 2), y=round(center_y, 2)),
                "status": DrawingResultRowStatus.FOUND,
                "note": "Точка поставлена по full-page truth; локальный OCR подтвердил метку в узком crop самой круговой выноски.",
                "source_kind": "vlm_locator_circle_confirmed",
                "final_score": max(float(_parse_number(truth_item.get("confidence")) or 0.0), float(local.confidence or 0.0)),
            }
        )

    vlm_recognizer = VisionLLMCandidateRecognizer()
    if not vlm_recognizer.is_enabled():
        return None
    vlm = vlm_recognizer.recognize(
        crop,
        "text",
        allowed_labels=[label],
        use_consensus=False,
        heavy_sheet=bool(min_side <= CANONICAL_MIN_SIDE),
    )
    if normalize_label(vlm.label) == normalized_label and (vlm.confidence or 0.0) >= 0.8:
        return _truth_only_row(truth_item, row_number=row_number).model_copy(
            update={
                "center": ResultPoint(x=round(center_x, 2), y=round(center_y, 2)),
                "top_left": ResultPoint(x=round(center_x, 2), y=round(center_y, 2)),
                "status": DrawingResultRowStatus.FOUND,
                "note": "Точка поставлена по full-page truth; локальная VLM-проверка подтвердила метку в узком crop самой круговой выноски.",
                "source_kind": "vlm_locator_circle_confirmed",
                "final_score": max(float(_parse_number(truth_item.get("confidence")) or 0.0), float(vlm.confidence or 0.0)),
            }
        )
    return None


def _refine_truth_only_row_from_local_crop(
    truth_item: Mapping[str, Any],
    *,
    raster_path: Path | None,
    page_width: int | None,
    page_height: int | None,
    row_number: int,
    require_local_match: bool,
) -> DrawingResultRow | None:
    if raster_path is None or not raster_path.is_file():
        return None

    label = str(truth_item.get("label") or "").strip()
    normalized_label = normalize_label(label)
    if not normalized_label:
        return None

    x = float(_parse_number(truth_item.get("x")) or 0.0)
    y = float(_parse_number(truth_item.get("y")) or 0.0)
    if x <= 0 or y <= 0:
        return None

    try:
        with Image.open(raster_path) as source_image:
            image = ImageOps.exif_transpose(source_image).convert("RGB")
    except Exception:
        return None
    local_recognizer = DrawingCandidateRecognizer()
    search_radii = [120, 160, 220]
    if len(label) >= 3:
        search_radii = [160, 220, 280]

    best_region = None
    best_region_score = None
    min_side = min(image.size)
    for crop_radius in search_radii:
        left = max(0, int(round(x - crop_radius)))
        top = max(0, int(round(y - crop_radius)))
        right = min(image.width, int(round(x + crop_radius)))
        bottom = min(image.height, int(round(y + crop_radius)))
        if right - left < 24 or bottom - top < 24:
            continue

        crop = image.crop((left, top, right, bottom))
        regions = local_recognizer.detect_document_text(crop, include_tiles=True)
        matched_regions: list[tuple[float, Any]] = []
        for region in regions:
            match_score = _targeted_label_match_score(region.label, label)
            if match_score <= 0.0:
                continue
            region_confidence = float(region.confidence or 0.0)
            if region_confidence < 0.7 and match_score < 1.0:
                continue
            matched_regions.append((match_score, region))
        if not matched_regions:
            continue

        for match_score, region in matched_regions:
            abs_center_x = left + region.bbox_x + region.bbox_width / 2.0
            abs_center_y = top + region.bbox_y + region.bbox_height / 2.0
            distance = hypot(abs_center_x - x, abs_center_y - y)
            score = float(region.confidence or 0.0) + (match_score - 1.0) * 0.2 - (distance / max(crop_radius * 6.0, 1.0))
            if best_region is None or best_region_score is None or score > best_region_score:
                best_region = (left, top, region, abs_center_x, abs_center_y, match_score)
                best_region_score = score
        if best_region is not None and best_region_score is not None and best_region_score >= 0.55:
            break

    if best_region is not None:
        left, top, region, abs_center_x, abs_center_y, match_score = best_region
        truth_row = _truth_only_row(truth_item, row_number=row_number)
        circle_anchor = None
        if match_score < 0.999:
            circle_anchor = _find_circle_callout_anchor(image, abs_center_x, abs_center_y)
        final_center_x = circle_anchor[0] if circle_anchor is not None else abs_center_x
        final_center_y = circle_anchor[1] if circle_anchor is not None else abs_center_y
        final_status = (
            DrawingResultRowStatus.FOUND
            if match_score >= 0.999 or circle_anchor is not None
            else DrawingResultRowStatus.UNCERTAIN
        )
        final_note = (
            "Точка поставлена по full-page truth и уточнена локальным OCR рядом с меткой."
            if match_score >= 0.999
            else (
                "Точка поставлена по full-page truth и уточнена локальным OCR с учётом типичной OCR-путаницы в суффиксе; круговая выноска подтверждена локально."
                if circle_anchor is not None
                else "Точка поставлена по full-page truth и уточнена локальным OCR с учётом типичной OCR-путаницы в суффиксе."
            )
        )
        return truth_row.model_copy(
            update={
                "center": ResultPoint(x=round(final_center_x, 2), y=round(final_center_y, 2)),
                "top_left": ResultPoint(x=round(left + region.bbox_x, 2), y=round(top + region.bbox_y, 2)),
                "bbox": ResultBoundingBox(
                    x=round(left + region.bbox_x, 2),
                    y=round(top + region.bbox_y, 2),
                    w=round(region.bbox_width, 2),
                    h=round(region.bbox_height, 2),
                ),
                "final_score": max(float(truth_row.final_score or 0.0), float(region.confidence or 0.0)),
                "status": final_status,
                "note": final_note,
                "source_kind": "vlm_locator_refined",
            }
        )

    if require_local_match:
        return None

    extended_exact_region = None
    extended_exact_score = None
    for crop_radius in [280, 340]:
        left = max(0, int(round(x - crop_radius)))
        top = max(0, int(round(y - crop_radius)))
        right = min(image.width, int(round(x + crop_radius)))
        bottom = min(image.height, int(round(y + crop_radius)))
        if right - left < 24 or bottom - top < 24:
            continue
        crop = image.crop((left, top, right, bottom))
        local = local_recognizer.recognize(crop, "text")
        vlm_recognizer = VisionLLMCandidateRecognizer()
        vlm_confirmed = False
        if normalize_label(local.label) == normalized_label and (local.confidence or 0.0) >= 0.72:
            vlm_confirmed = True
        elif vlm_recognizer.is_enabled():
            vlm = vlm_recognizer.recognize(
                crop,
                "text",
                allowed_labels=[label],
                use_consensus=False,
                heavy_sheet=bool(min_side <= CANONICAL_MIN_SIDE),
            )
            vlm_confirmed = normalize_label(vlm.label) == normalized_label and (vlm.confidence or 0.0) >= 0.82
        if not vlm_confirmed:
            continue

        regions = local_recognizer.detect_document_text(crop, include_tiles=True)
        for region in regions:
            if _targeted_label_match_score(region.label, label) < 0.999:
                continue
            if float(region.confidence or 0.0) < 0.72:
                continue
            abs_center_x = left + region.bbox_x + region.bbox_width / 2.0
            abs_center_y = top + region.bbox_y + region.bbox_height / 2.0
            distance = hypot(abs_center_x - x, abs_center_y - y)
            score = float(region.confidence or 0.0) - (distance / max(crop_radius * 7.0, 1.0))
            if extended_exact_region is None or extended_exact_score is None or score > extended_exact_score:
                extended_exact_region = (left, top, region, abs_center_x, abs_center_y)
                extended_exact_score = score
        if extended_exact_region is not None:
            left, top, region, abs_center_x, abs_center_y = extended_exact_region
            truth_row = _truth_only_row(truth_item, row_number=row_number)
            return truth_row.model_copy(
                update={
                    "center": ResultPoint(x=round(abs_center_x, 2), y=round(abs_center_y, 2)),
                    "top_left": ResultPoint(x=round(left + region.bbox_x, 2), y=round(top + region.bbox_y, 2)),
                    "bbox": ResultBoundingBox(
                        x=round(left + region.bbox_x, 2),
                        y=round(top + region.bbox_y, 2),
                        w=round(region.bbox_width, 2),
                        h=round(region.bbox_height, 2),
                    ),
                    "final_score": max(float(_parse_number(truth_item.get("confidence")) or 0.0), float(region.confidence or 0.0)),
                    "status": DrawingResultRowStatus.FOUND,
                    "note": "Точка поставлена по full-page truth и уточнена локальным OCR в расширенном локальном окне.",
                    "source_kind": "vlm_locator_refined",
                }
            )

    left = max(0, int(round(x - 140)))
    top = max(0, int(round(y - 140)))
    right = min(image.width, int(round(x + 140)))
    bottom = min(image.height, int(round(y + 140)))
    if right - left < 24 or bottom - top < 24:
        return None

    crop = image.crop((left, top, right, bottom))
    local = local_recognizer.recognize(crop, "text")
    if normalize_label(local.label) == normalized_label and (local.confidence or 0.0) >= 0.72:
        circle_anchor = _find_circle_callout_anchor(image, x, y)
        updates: dict[str, Any] = {
            "note": "Точка поставлена по full-page truth; локальный OCR подтвердил метку рядом, но не дал bbox.",
            "source_kind": "vlm_locator_confirmed",
            "final_score": max(float(_parse_number(truth_item.get('confidence')) or 0.0), float(local.confidence or 0.0)),
        }
        if circle_anchor is not None:
            updates.update(
                {
                    "center": ResultPoint(x=round(circle_anchor[0], 2), y=round(circle_anchor[1], 2)),
                    "top_left": ResultPoint(x=round(circle_anchor[0], 2), y=round(circle_anchor[1], 2)),
                    "status": DrawingResultRowStatus.FOUND,
                    "note": "Точка поставлена по full-page truth; локальный OCR подтвердил метку рядом, круговая выноска подтверждена локально.",
                    "source_kind": "vlm_locator_circle_confirmed",
                }
            )
        return _truth_only_row(truth_item, row_number=row_number).model_copy(update=updates)

    vlm_recognizer = VisionLLMCandidateRecognizer()
    if vlm_recognizer.is_enabled():
        vlm = vlm_recognizer.recognize(
            crop,
            "text",
            allowed_labels=[label],
            use_consensus=False,
            heavy_sheet=bool(min_side <= CANONICAL_MIN_SIDE),
        )
        if normalize_label(vlm.label) == normalized_label and (vlm.confidence or 0.0) >= 0.82:
            circle_anchor = _find_circle_callout_anchor(image, x, y)
            updates = {
                "note": "Точка поставлена по full-page truth; локальная VLM-проверка подтвердила метку рядом.",
                "source_kind": "vlm_locator_confirmed",
                "final_score": max(float(_parse_number(truth_item.get('confidence')) or 0.0), float(vlm.confidence or 0.0)),
            }
            if circle_anchor is not None:
                updates.update(
                    {
                        "center": ResultPoint(x=round(circle_anchor[0], 2), y=round(circle_anchor[1], 2)),
                        "top_left": ResultPoint(x=round(circle_anchor[0], 2), y=round(circle_anchor[1], 2)),
                        "status": DrawingResultRowStatus.FOUND,
                        "note": "Точка поставлена по full-page truth; локальная VLM-проверка подтвердила метку рядом, круговая выноска подтверждена локально.",
                        "source_kind": "vlm_locator_circle_confirmed",
                    }
                )
            return _truth_only_row(truth_item, row_number=row_number).model_copy(update=updates)
    circle_anchor = _find_circle_callout_anchor(image, x, y)
    if circle_anchor is not None:
        circle_row = _refine_truth_only_row_from_circle_crop(
            truth_item,
            image=image,
            label=label,
            normalized_label=normalized_label,
            center_x=circle_anchor[0],
            center_y=circle_anchor[1],
            radius=circle_anchor[2],
            row_number=row_number,
            min_side=min_side,
            local_recognizer=local_recognizer,
        )
        if circle_row is not None:
            return circle_row
        if circle_anchor[3] >= 1.55:
            return _truth_only_row(truth_item, row_number=row_number).model_copy(
                update={
                    "center": ResultPoint(x=round(circle_anchor[0], 2), y=round(circle_anchor[1], 2)),
                    "top_left": ResultPoint(x=round(circle_anchor[0] - circle_anchor[2], 2), y=round(circle_anchor[1] - circle_anchor[2], 2)),
                    "bbox": ResultBoundingBox(
                        x=round(circle_anchor[0] - circle_anchor[2], 2),
                        y=round(circle_anchor[1] - circle_anchor[2], 2),
                        w=round(circle_anchor[2] * 2.0, 2),
                        h=round(circle_anchor[2] * 2.0, 2),
                    ),
                    "status": DrawingResultRowStatus.FOUND,
                    "note": "Точка поставлена по full-page truth; очень сильная геометрия круговой выноски подтверждена локально.",
                    "source_kind": "vlm_locator_circle_confirmed",
                    "final_score": max(float(_parse_number(truth_item.get('confidence')) or 0.0), min(0.98, 0.78 + circle_anchor[3] * 0.1)),
                }
            )
    return None


def _suppress_redundant_truth_only_rows(
    rows: list[DrawingResultRow],
    held_back_rows: list[DrawingResultRow],
    *,
    page_sizes_by_index: Mapping[int, tuple[int, int]] | None,
) -> tuple[list[DrawingResultRow], list[DrawingResultRow]]:
    if not rows:
        return rows, held_back_rows

    kept_rows: list[DrawingResultRow] = []
    extra_held_back = list(held_back_rows)
    by_page_and_label: dict[tuple[int, str], list[DrawingResultRow]] = {}
    for row in rows:
        normalized = normalize_label(row.label)
        if not normalized:
            kept_rows.append(row)
            continue
        by_page_and_label.setdefault((row.page_index, normalized), []).append(row)

    for group_rows in by_page_and_label.values():
        stable_rows = [
            row
            for row in group_rows
            if not (row.status == DrawingResultRowStatus.UNCERTAIN and row.source_kind == "vlm_locator")
        ]
        truth_only_rows = [
            row
            for row in group_rows
            if row.status == DrawingResultRowStatus.UNCERTAIN and row.source_kind == "vlm_locator"
        ]
        if not truth_only_rows:
            kept_rows.extend(group_rows)
            continue

        page_width, page_height = (page_sizes_by_index or {}).get(group_rows[0].page_index, (None, None))
        duplicate_radius = _truth_only_duplicate_radius_px(page_width, page_height)
        kept_group = list(stable_rows)
        for truth_row in truth_only_rows:
            truth_center = truth_row.center
            if truth_center is None:
                extra_held_back.append(
                    truth_row.model_copy(
                        update={
                            "note": "Кандидат удержан: full-page truth дал дублирующую точку без локального подтверждения."
                        }
                    )
                )
                continue

            duplicate_peer = next(
                (
                    peer
                    for peer in stable_rows
                    if peer.center is not None
                    and hypot(peer.center.x - truth_center.x, peer.center.y - truth_center.y) <= duplicate_radius
                ),
                None,
            )
            if duplicate_peer is not None:
                extra_held_back.append(
                    truth_row.model_copy(
                        update={
                            "note": (
                                "Кандидат удержан: full-page truth дал вторую точку слишком близко к уже найденной "
                                "локальной метке."
                            )
                        }
                    )
                )
                continue
            kept_group.append(truth_row)
        kept_rows.extend(kept_group)

    kept_rows.sort(key=lambda row: (row.page_index, row.row, row.label))
    extra_held_back.sort(key=lambda row: (row.page_index, row.row, row.label))
    return kept_rows, extra_held_back
def _build_pipeline_failure_result(
    *,
    job: DrawingJob,
    prepared_pages: list[PreparedDrawingPage],
    selected_ocr_engine: str,
    fallback_used: bool,
    failure_message: str,
) -> DrawingJobResult:
    pages = [
        DrawingResultPage(
            page_index=page.page_index,
            width=page.width,
            height=page.height,
            overlay_url=None,
            row_count=0,
            held_back_count=0,
        )
        for page in prepared_pages
    ]
    review_reason = f"Первичный OCR-проход завершился ошибкой: {failure_message}"
    summary = DrawingJobSummary(
        total_rows=0,
        found_count=0,
        missing_count=0,
        uncertain_count=0,
        held_back_count=0,
        discarded_count=0,
        document_confidence=0.0,
        degraded_recognition=not job.input.has_labels,
        degraded_reason="Первичный OCR-проход не дал результата. Ищу более дешёвый запасной маршрут." if not job.input.has_labels else None,
        selected_ocr_engine=selected_ocr_engine,
        fallback_used=fallback_used,
        review_recommended=True,
        review_reasons=[review_reason],
        status_text="Первичный OCR-проход не дал результата. Пробую запасной маршрут." if not job.input.has_labels else failure_message,
        failure_message=failure_message,
    )
    return DrawingJobResult(
        source_file=job.input.drawing_name,
        source_labels_file=job.input.labels_name,
        pages=pages,
        rows=[],
        held_back_rows=[],
        missing_labels=[],
        extra_detected_labels=[],
        summary=summary,
    )


def _promote_emergency_review_rows(result: DrawingJobResult) -> int:
    promotable: list[DrawingResultRow] = []
    remaining: list[DrawingResultRow] = []
    for row in result.held_back_rows:
        score = row.final_score or 0.0
        note = row.note or ""
        if NEAR_TIE_NOTE_TOKEN in note:
            remaining.append(row)
            continue
        if row.source_kind == "horizontal" and score >= 0.54:
            promotable.append(
                row.model_copy(
                    update={
                        "status": DrawingResultRowStatus.UNCERTAIN,
                        "note": (
                            ((row.note or "").strip() + " ") if row.note else ""
                        )
                        + "Поднят из held-back в review, потому что аварийный OCR-маршрут не нашёл надёжного результата."
                    }
                )
            )
        else:
            remaining.append(row)

    if not promotable:
        return 0

    existing_keys = {(row.label, row.page_index) for row in result.rows}
    promoted_count = 0
    for row in promotable:
        key = (row.label, row.page_index)
        if key in existing_keys:
            continue
        result.rows.append(row)
        existing_keys.add(key)
        promoted_count += 1

    result.rows.sort(key=lambda row: (row.page_index, row.row, row.label))
    result.held_back_rows = remaining

    row_count_by_page: dict[int, int] = {}
    held_back_count_by_page: dict[int, int] = {}
    for row in result.rows:
        if row.status == DrawingResultRowStatus.NOT_FOUND:
            continue
        row_count_by_page[row.page_index] = row_count_by_page.get(row.page_index, 0) + 1
    for row in result.held_back_rows:
        held_back_count_by_page[row.page_index] = held_back_count_by_page.get(row.page_index, 0) + 1
    for page in result.pages:
        page.row_count = row_count_by_page.get(page.page_index, 0)
        page.held_back_count = held_back_count_by_page.get(page.page_index, 0)
    return promoted_count


def _should_try_degraded_fallback(job: DrawingJob, result: DrawingJobResult) -> bool:
    return (
        not job.input.has_labels
        and (
            result.summary.failure_message is not None
            or (
                result.summary.degraded_recognition
                and (result.summary.document_confidence or 0.0) < 0.72
            )
        )
    )


def _load_combined_source_markers(source_json_path: Path) -> list[dict[str, Any]]:
    if not source_json_path.exists():
        return []
    try:
        payload = json.loads(source_json_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    pages = payload.get("pages")
    if not isinstance(pages, list):
        return []
    markers: list[dict[str, Any]] = []
    for page in pages:
        if not isinstance(page, Mapping):
            continue
        page_index = int(_parse_number(page.get("page_index")) or 0)
        page_payload = page.get("payload")
        if not isinstance(page_payload, Mapping):
            continue
        for item in page_payload.get("markers", []):
            if not isinstance(item, Mapping):
                continue
            marker = dict(item)
            marker["_page_index"] = page_index
            markers.append(marker)
    return markers


def _bbox_match_key(page_index: int, bbox: ResultBoundingBox | Mapping[str, Any] | None) -> tuple[int, int, int, int, int] | None:
    if bbox is None:
        return None
    if isinstance(bbox, ResultBoundingBox):
        x = bbox.x
        y = bbox.y
        w = bbox.w
        h = bbox.h
    elif isinstance(bbox, Mapping):
        x = _parse_number(bbox.get("x"))
        y = _parse_number(bbox.get("y"))
        w = _parse_number(bbox.get("w"))
        h = _parse_number(bbox.get("h"))
    else:
        return None
    if x is None or y is None or w is None or h is None:
        return None
    return (page_index, round(x * 10), round(y * 10), round(w * 10), round(h * 10))


def _score_map_from_marker(marker: Mapping[str, Any]) -> list[tuple[str, float]]:
    recognition = marker.get("recognition")
    if not isinstance(recognition, Mapping):
        return []
    merged: dict[str, float] = {}
    for key in ("easyocr_scores", "rapidocr_scores"):
        raw_scores = recognition.get(key)
        if not isinstance(raw_scores, Mapping):
            continue
        for raw_label, raw_score in raw_scores.items():
            label = str(raw_label or "").strip()
            score = _parse_number(raw_score)
            if not label or score is None:
                continue
            existing = merged.get(label)
            if existing is None or score > existing:
                merged[label] = score
    if not merged:
        best_label = str(recognition.get("ocr_best_label") or marker.get("label") or "").strip()
        best_score = _parse_number(recognition.get("ocr_best_score"))
        if best_label and best_score is not None:
            merged[best_label] = best_score
    return sorted(merged.items(), key=lambda item: (-item[1], item[0]))


def _collect_near_tie_export_rows(
    result: DrawingJobResult,
    *,
    source_json_path: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    near_tie_rows = [row for row in result.held_back_rows if row.note and NEAR_TIE_NOTE_TOKEN in row.note]
    if not near_tie_rows:
        return [], []

    combined_markers = _load_combined_source_markers(source_json_path)
    markers_by_key: dict[tuple[int, int, int, int, int], list[dict[str, Any]]] = {}
    for marker in combined_markers:
        page_index = int(_parse_number(marker.get("_page_index")) or 0)
        key = _bbox_match_key(page_index, marker.get("bbox") if isinstance(marker.get("bbox"), Mapping) else None)
        if key is None:
            continue
        markers_by_key.setdefault(key, []).append(marker)

    csv_rows: list[dict[str, Any]] = []
    json_rows: list[dict[str, Any]] = []
    for row in near_tie_rows:
        key = _bbox_match_key(row.page_index, row.bbox)
        source_marker: dict[str, Any] | None = None
        if key is not None:
            bucket = markers_by_key.get(key) or []
            if bucket:
                source_marker = bucket.pop(0)

        ranked_scores = _score_map_from_marker(source_marker or {})
        recognition = source_marker.get("recognition") if isinstance(source_marker, Mapping) else None
        ocr_best_label = str((recognition or {}).get("ocr_best_label") or row.label).strip() or row.label
        ocr_best_score = _parse_number((recognition or {}).get("ocr_best_score"))
        ocr_second_score = _parse_number((recognition or {}).get("ocr_second_score"))
        alternative_label = ""
        if len(ranked_scores) >= 2:
            alternative_label = ranked_scores[1][0]
            if ocr_best_score is None:
                ocr_best_score = ranked_scores[0][1]
            if ocr_second_score is None:
                ocr_second_score = ranked_scores[1][1]
        ocr_gap = _parse_number((source_marker or {}).get("ocr_gap"))
        if ocr_gap is None and ocr_best_score is not None and ocr_second_score is not None:
            ocr_gap = ocr_best_score - ocr_second_score

        csv_row = {
            "row": row.row,
            "page_index": row.page_index,
            "label": row.label,
            "alternative_label": alternative_label,
            "ocr_best_label": ocr_best_label,
            "ocr_best_score": ocr_best_score,
            "ocr_second_score": ocr_second_score,
            "ocr_gap": ocr_gap,
            "center_x": row.center.x if row.center else None,
            "center_y": row.center.y if row.center else None,
            "bbox_x": row.bbox.x if row.bbox else None,
            "bbox_y": row.bbox.y if row.bbox else None,
            "bbox_w": row.bbox.w if row.bbox else None,
            "bbox_h": row.bbox.h if row.bbox else None,
            "source_kind": row.source_kind or "",
            "note": row.note or "",
        }
        csv_rows.append(csv_row)
        json_rows.append(
            {
                **csv_row,
                "ranked_scores": [{"label": label, "score": score} for label, score in ranked_scores],
            }
        )
    return csv_rows, json_rows


def _should_replace_primary_result(primary: DrawingJobResult, fallback: DrawingJobResult) -> bool:
    primary_summary = primary.summary
    fallback_summary = fallback.summary

    if fallback_summary.failure_message:
        return False
    if fallback_summary.found_count < primary_summary.found_count:
        return False
    if fallback_summary.uncertain_count > primary_summary.uncertain_count:
        return False
    if fallback_summary.held_back_count > primary_summary.held_back_count:
        return False

    primary_confidence = primary_summary.document_confidence or 0.0
    fallback_confidence = fallback_summary.document_confidence or 0.0

    if primary_summary.failure_message and not fallback_summary.failure_message:
        return True
    if primary_summary.degraded_recognition and not fallback_summary.degraded_recognition:
        return True
    if fallback_confidence >= primary_confidence + 0.10:
        return True
    return False


def _fallback_cli_args() -> list[str]:
    args = [
        "--detect-scales",
        settings.legacy_fallback_detect_scales,
        "--tile-size",
        str(settings.legacy_fallback_tile_size),
        "--tile-overlap",
        str(settings.legacy_fallback_tile_overlap),
    ]
    if settings.legacy_fallback_disable_gemini:
        args.append("--disable-gemini")
    if settings.legacy_fallback_disable_gemini_tile_proposals:
        args.append("--disable-gemini-tile-proposals")
    return args


def _emergency_fallback_cli_args() -> list[str]:
    return [
        "--detect-scales",
        "1.0",
        "--tile-size",
        "1536",
        "--tile-overlap",
        "96",
        "--disable-gemini",
        "--disable-gemini-tile-proposals",
    ]


def run_job_pipeline(job_dir: Path, job: DrawingJob) -> JobRunOutput:
    input_dir = job_dir / "input"
    work_dir = job_dir / "work"
    pipeline_dir = work_dir / "pipeline"
    logs_dir = job_dir / "logs"
    artifacts_dir = job_dir / "artifacts"
    work_dir.mkdir(parents=True, exist_ok=True)
    artifacts_dir.mkdir(parents=True, exist_ok=True)

    drawing_path = input_dir / f"drawing{Path(job.input.drawing_name).suffix}"
    if not drawing_path.is_file():
        raise RuntimeError(f"Файл чертежа не найден: {drawing_path.name}")

    labels_path: Path | None = None
    if job.input.labels_name:
        labels_path = input_dir / f"labels{Path(job.input.labels_name).suffix}"
        if not labels_path.is_file():
            raise RuntimeError(f"Файл таблицы не найден: {labels_path.name}")

    prepared_drawing_path = work_dir / "prepared-drawing.png"
    prepared_pages = prepare_drawing_for_legacy_pipeline(drawing_path, prepared_drawing_path)

    labels_xlsx_path, expected_labels = prepare_labels_for_legacy_pipeline(labels_path, work_dir)
    primary_engine = _normalize_ocr_engine(os.getenv("WEBUI_OCR_ENGINE", "both"))

    def execute_pipeline_pass(
        *,
        ocr_engine: str,
        pass_slug: str,
        fallback_used: bool,
        timeout_seconds: int | None = None,
        extra_cli_args: list[str] | None = None,
    ) -> tuple[DrawingJobResult, list[JobPageArtifact], Path]:
        pass_pipeline_dir = pipeline_dir / pass_slug
        pass_logs_dir = logs_dir / pass_slug
        pass_artifacts_dir = artifacts_dir / pass_slug
        pass_artifacts_dir.mkdir(parents=True, exist_ok=True)

        page_payloads: list[dict[str, Any]] = []
        page_artifacts: list[JobPageArtifact] = []
        combined_source_payload: dict[str, Any] = {
            "source_file": job.input.drawing_name,
            "source_labels_file": job.input.labels_name,
            "selected_ocr_engine": ocr_engine,
            "fallback_used": fallback_used,
            "pages": [],
        }

        for prepared_page in prepared_pages:
            page_slug = f"page-{prepared_page.page_index + 1:03d}"
            legacy_output = run_legacy_pipeline(
                image_path=prepared_page.raster_path,
                labels_xlsx_path=labels_xlsx_path,
                out_dir=pass_pipeline_dir / page_slug,
                log_dir=pass_logs_dir / page_slug,
                ocr_engine=ocr_engine,
                timeout_seconds=timeout_seconds,
                extra_cli_args=extra_cli_args,
            )
            page_payloads.append(legacy_output.payload)

            overlay_artifact_path: Path | None = None
            if legacy_output.overlay_path.exists():
                overlay_artifact_path = pass_artifacts_dir / f"{page_slug}.overlay.png"
                shutil.copy2(legacy_output.overlay_path, overlay_artifact_path)

            page_source_json_artifact_path: Path | None = None
            if legacy_output.markers_json_path.exists():
                page_source_json_artifact_path = pass_artifacts_dir / f"{page_slug}.markers.json"
                shutil.copy2(legacy_output.markers_json_path, page_source_json_artifact_path)

            page_artifacts.append(
                JobPageArtifact(
                    page_index=prepared_page.page_index,
                    overlay_path=overlay_artifact_path,
                    source_json_path=page_source_json_artifact_path,
                    width=prepared_page.width,
                    height=prepared_page.height,
                    raster_path=prepared_page.raster_path,
                )
            )
            combined_source_payload["pages"].append(
                {
                    "page_index": prepared_page.page_index,
                    "width": prepared_page.width,
                    "height": prepared_page.height,
                    "markers_path": page_source_json_artifact_path.name if page_source_json_artifact_path else None,
                    "report": legacy_output.report,
                    "payload": legacy_output.payload,
                }
            )

        source_json_path = pass_artifacts_dir / "source-markers.json"
        source_json_path.write_text(json.dumps(combined_source_payload, ensure_ascii=False, indent=2), encoding="utf-8")

        result = build_result_from_legacy_output(
            job=job,
            page_artifacts=page_artifacts,
            page_payloads=page_payloads,
            expected_labels=expected_labels,
            selected_ocr_engine=ocr_engine,
            fallback_used=fallback_used,
        )
        return result, page_artifacts, source_json_path

    try:
        result, page_artifacts, source_json_artifact_path = execute_pipeline_pass(
            ocr_engine=primary_engine,
            pass_slug="primary",
            fallback_used=False,
            timeout_seconds=settings.legacy_pipeline_timeout_seconds,
        )
    except Exception as primary_error:
        if job.input.has_labels:
            raise
        primary_message = str(primary_error)
        (logs_dir / "primary.error.txt").write_text(primary_message, encoding="utf-8")
        page_artifacts = [
            JobPageArtifact(
                page_index=page.page_index,
                overlay_path=None,
                source_json_path=None,
                width=page.width,
                height=page.height,
                raster_path=page.raster_path,
            )
            for page in prepared_pages
        ]
        source_json_artifact_path = artifacts_dir / "source-markers.json"
        result = _build_pipeline_failure_result(
            job=job,
            prepared_pages=prepared_pages,
            selected_ocr_engine=primary_engine,
            fallback_used=False,
            failure_message=primary_message,
        )

    fallback_failures: list[str] = []
    fallback_attempted = False
    emergency_fallback_used = False
    emergency_fallback_reason: str | None = None
    if _should_try_degraded_fallback(job, result):
        fallback_attempted = True
        selected_bundle = (result, page_artifacts, source_json_artifact_path)
        selected_result = result
        for fallback_engine in _fallback_ocr_engines(primary_engine):
            try:
                fallback_result, fallback_page_artifacts, fallback_source_json_path = execute_pipeline_pass(
                    ocr_engine=fallback_engine,
                    pass_slug=f"fallback-{fallback_engine}",
                    fallback_used=True,
                    timeout_seconds=settings.legacy_fallback_pipeline_timeout_seconds,
                    extra_cli_args=_fallback_cli_args(),
                )
            except Exception as fallback_error:
                message = str(fallback_error)
                (logs_dir / f"fallback-{fallback_engine}.error.txt").write_text(message, encoding="utf-8")
                fallback_failures.append(f"{fallback_engine}: {message}")
                continue
            if _should_replace_primary_result(selected_result, fallback_result):
                selected_bundle = (fallback_result, fallback_page_artifacts, fallback_source_json_path)
                selected_result = fallback_result
        if (
            settings.legacy_emergency_fallback_enabled
            and selected_result.summary.failure_message is not None
            and primary_engine != "rapid"
        ):
            try:
                emergency_result, emergency_page_artifacts, emergency_source_json_path = execute_pipeline_pass(
                    ocr_engine="rapid",
                    pass_slug="emergency-fallback-rapid",
                    fallback_used=True,
                    timeout_seconds=settings.legacy_emergency_fallback_pipeline_timeout_seconds,
                    extra_cli_args=_emergency_fallback_cli_args(),
                )
            except Exception as emergency_error:
                message = str(emergency_error)
                (logs_dir / "emergency-fallback-rapid.error.txt").write_text(message, encoding="utf-8")
                fallback_failures.append(f"emergency-rapid: {message}")
            else:
                if emergency_result.summary.found_count > 0 or emergency_result.summary.uncertain_count > 0:
                    selected_bundle = (emergency_result, emergency_page_artifacts, emergency_source_json_path)
                    selected_result = emergency_result
                    emergency_fallback_used = True
                    emergency_fallback_reason = (
                        "Основной OCR-проход не дал результата, и аварийный rapid fallback собрал черновой результат для review."
                    )
        result, page_artifacts, source_json_artifact_path = selected_bundle

    result.summary.fallback_attempted = fallback_attempted
    result.summary.fallback_failure_count = len(fallback_failures)
    result.summary.emergency_fallback_used = emergency_fallback_used
    result.summary.emergency_fallback_reason = emergency_fallback_reason
    if fallback_failures:
        result.summary.review_reasons.append(
            f"Запасной OCR-проход пытался помочь, но {len(fallback_failures)} вариант(ов) завершились ошибкой или timeout."
        )
    if emergency_fallback_reason:
        promoted_count = _promote_emergency_review_rows(result)
        result.summary.degraded_recognition = True
        result.summary.degraded_reason = result.summary.degraded_reason or emergency_fallback_reason
        result.summary.failure_message = None
        result.summary.review_recommended = True
        result.summary.status_text = "Основной OCR не справился, но аварийный маршрут собрал черновой результат для ручной проверки."
        result.summary.total_rows = len(result.rows)
        result.summary.found_count = sum(1 for row in result.rows if row.status == DrawingResultRowStatus.FOUND)
        result.summary.missing_count = sum(1 for row in result.rows if row.status == DrawingResultRowStatus.NOT_FOUND)
        result.summary.uncertain_count = sum(1 for row in result.rows if row.status == DrawingResultRowStatus.UNCERTAIN)
        result.summary.held_back_count = len(result.held_back_rows)
        result.summary.near_tie_ambiguity_count = sum(
            1 for row in result.held_back_rows if row.note and NEAR_TIE_NOTE_TOKEN in row.note
        )
        result.summary.review_reasons = [
            reason
            for reason in result.summary.review_reasons
            if "точек ниже порога уверенности 70%" not in reason
            and "слабых кандидатов удержаны" not in reason
        ]
        if result.summary.uncertain_count > 0:
            result.summary.review_reasons.insert(
                0,
                f"{result.summary.uncertain_count} точек ниже порога уверенности 70%.",
            )
        if result.summary.held_back_count > 0:
            result.summary.review_reasons.append(
                f"{result.summary.held_back_count} слабых кандидатов удержаны для ручной проверки и не попали в итоговый список."
            )
        if promoted_count > 0:
            result.summary.review_reasons.append(
                f"{promoted_count} удержанных horizontal OCR-only кандидатов подняты в review-слой, потому что аварийный OCR не нашёл надёжного результата."
            )
        result.summary.review_reasons.append(emergency_fallback_reason)

    for artifact in page_artifacts:
        if artifact.overlay_path is not None and artifact.overlay_path.exists():
            shutil.copy2(artifact.overlay_path, artifacts_dir / artifact.overlay_path.name)
        if artifact.source_json_path is not None and artifact.source_json_path.exists():
            shutil.copy2(artifact.source_json_path, artifacts_dir / artifact.source_json_path.name)
    if source_json_artifact_path.exists():
        shutil.copy2(source_json_artifact_path, artifacts_dir / "source-markers.json")
    source_json_artifact_path = artifacts_dir / "source-markers.json"

    stem = safe_slug(Path(job.input.drawing_name).stem or job.title or "coordinates", default="coordinates")
    production_rows = [
        row for row in result.rows if row.status in {DrawingResultRowStatus.FOUND, DrawingResultRowStatus.NOT_FOUND}
    ]
    production_csv_path = write_result_csv(result, artifacts_dir / f"{stem}.csv", rows=production_rows)
    production_xlsx_path = write_result_xlsx(
        result,
        artifacts_dir / f"{stem}.xlsx",
        rows=production_rows,
        export_label="production_without_uncertain",
    )
    review_csv_path = write_result_csv(result, artifacts_dir / f"{stem}.review.csv")
    review_xlsx_path = write_result_xlsx(
        result,
        artifacts_dir / f"{stem}.review.xlsx",
        export_label="review_with_uncertain",
    )
    held_back_csv_path = None
    near_tie_csv_path = None
    near_tie_json_path = None
    if result.held_back_rows:
        held_back_csv_path = write_result_csv(
            result,
            artifacts_dir / f"{stem}.held-back.csv",
            rows=result.held_back_rows,
        )
        near_tie_csv_rows, near_tie_json_rows = _collect_near_tie_export_rows(
            result,
            source_json_path=source_json_artifact_path,
        )
        if near_tie_csv_rows:
            near_tie_columns = [
                "row",
                "page_index",
                "label",
                "alternative_label",
                "ocr_best_label",
                "ocr_best_score",
                "ocr_second_score",
                "ocr_gap",
                "center_x",
                "center_y",
                "bbox_x",
                "bbox_y",
                "bbox_w",
                "bbox_h",
                "source_kind",
                "note",
            ]
            near_tie_csv_path = write_table_csv(
                artifacts_dir / f"{stem}.near-tie.csv",
                columns=near_tie_columns,
                rows=near_tie_csv_rows,
            )
            near_tie_json_path = write_json_payload(
                {
                    "source_file": result.source_file,
                    "count": len(near_tie_json_rows),
                    "items": near_tie_json_rows,
                },
                artifacts_dir / f"{stem}.near-tie.json",
            )
    result_json_path = write_result_json(result, artifacts_dir / "result.json")
    production_zip_inputs = [production_csv_path, production_xlsx_path]
    production_zip_inputs.extend(artifact.overlay_path for artifact in page_artifacts if artifact.overlay_path is not None)
    production_zip_path = write_result_zip(artifacts_dir / f"{stem}.zip", production_zip_inputs)
    review_zip_inputs = [review_csv_path, review_xlsx_path, result_json_path, source_json_artifact_path]
    if held_back_csv_path is not None:
        review_zip_inputs.append(held_back_csv_path)
    if near_tie_csv_path is not None:
        review_zip_inputs.append(near_tie_csv_path)
    if near_tie_json_path is not None:
        review_zip_inputs.append(near_tie_json_path)
    review_zip_inputs.extend(artifact.overlay_path for artifact in page_artifacts if artifact.overlay_path is not None)
    review_zip_inputs.extend(artifact.source_json_path for artifact in page_artifacts if artifact.source_json_path is not None)
    review_zip_path = write_result_zip(artifacts_dir / f"{stem}.review.zip", review_zip_inputs)

    return JobRunOutput(
        result=result,
        production_csv_path=production_csv_path,
        production_xlsx_path=production_xlsx_path,
        production_zip_path=production_zip_path,
        review_csv_path=review_csv_path,
        held_back_csv_path=held_back_csv_path,
        near_tie_csv_path=near_tie_csv_path,
        near_tie_json_path=near_tie_json_path,
        review_xlsx_path=review_xlsx_path,
        review_zip_path=review_zip_path,
        source_json_path=source_json_artifact_path,
        result_json_path=result_json_path,
        page_artifacts=page_artifacts,
    )
